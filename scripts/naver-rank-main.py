"""
네이버 블로그 순위 확인 도구
- 키워드 목록 + 타깃 블로그 ID 입력
- 아래 세 곳에서만 블로그 순위 확인:
  1) 통합검색 → 검색결과 섹션
  2) 통합검색 → 반려동물 인기글 섹션
  3) 통합검색 상단 → 블로그 탭 (블로그 전용 페이지)
- 플레이스(스마트플레이스) 순위: 통합검색 플레이스 섹션에서 상호명으로 순위 확인 (광고 제외, 2·3페이지까지)

엑셀 입·출력 (여러 블로그 × 각자 키워드 한 번에):
- 입력 파일 (input.xlsx) — 두 종류 시트 사용 가능:

  [시트1] 블로그 순위용 (첫 시트 / 활성 시트)
    - 1행: A열 "블로그 ID", B열 "키워드" (헤더)
    - 2행부터: 한 행 = (블로그 ID, 확인할 키워드). 블로그 ID 비우면 위 행 값 유지
    예시:
      |   A (블로그 ID)  |   B (키워드)    |
      |------------------|-----------------|
    1 | 블로그 ID        | 키워드          |
    2 | jd_ah            | 은평구동물병원   |
    3 | jd_ah            | 수색동물병원     |
    4 |                  | 강아지미용       |  ← jd_ah와 동일

  [시트2] 플레이스 순위용 — 시트 이름 반드시 "플레이스"
    - 1행: A열 "키워드", B열 "상호명" (헤더)
    - 2행부터: 한 행 = (검색할 키워드, 찾을 매장 상호명). 둘 다 필수
    예시:
      |   A (키워드)   |   B (상호명)     |
      |----------------|------------------|
    1 | 키워드         | 상호명           |
    2 | 마포 동물병원  | 힐링동물병원 & 건강검진센터 |
    3 | 은평구 동물병원 | 정든동물병원     |

  - 블로그만 쓰려면 첫 시트만 채우면 됨. 플레이스만 쓰려면 "플레이스" 시트만 채우면 됨. 둘 다 채우면 두 순위 모두 실행.
  - input.xlsx가 없으면 main.py 상단 TARGET_BLOG_ID, KEYWORDS만 사용 (블로그만).

- 출력 (output.xlsx):
  - 시트 "순위결과": 블로그 순위 (블로그 ID, 키워드, 검색결과, 반려동물 인기글, 일반 검색, 블로그(탭), 노출URL, 비고)
  - 시트 "플레이스 순위": 플레이스 순위 (키워드, 상호명, 순위, 비고) — 입력에 "플레이스" 시트가 있을 때만 생성

- 실행: python main.py [입력엑셀] [출력엑셀]
"""

import asyncio
import time
import random
import re
import os
import json
import sys
import threading
import urllib.request
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from contextlib import suppress
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote, urlparse, parse_qs, unquote, urlencode
from urllib.request import Request, urlopen
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl


def load_local_env_files():
    """
    프로젝트 루트(.env/.env.local)의 KEY=VALUE를 현재 프로세스 환경으로 로드한다.
    이미 설정된 환경변수는 덮어쓰지 않는다.
    """
    project_root = Path(__file__).resolve().parents[1]
    candidates = [project_root / ".env", project_root / ".env.local"]

    for env_path in candidates:
        if not env_path.exists():
            continue
        try:
            for raw_line in env_path.read_text(encoding="utf-8").splitlines():
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip("'").strip('"')
                if key:
                    os.environ.setdefault(key, value)
        except Exception:
            # 환경 파일 파싱 실패는 치명적이지 않으므로 무시하고 기존 환경변수만 사용
            pass


load_local_env_files()

# 타깃 블로그 ID (예: blog.naver.com/여기부분)
TARGET_BLOG_ID = "jd_ah"

# 확인할 키워드 목록
KEYWORDS = [
    "은평구동물병원",
    "수색동물병원",
    "상암동동물병원",
    "가좌동동물병원"
]

SECTION_SPECS = [
    ("검색결과", "검색결과_URL", "검색결과", "blog_rank_integrated"),
    ("반려동물 인기글", "반려동물 인기글_URL", "반려동물 인기글", "blog_rank_pet_popular"),
    ("일반 검색", "일반 검색_URL", "일반 검색", "blog_rank_general"),
    ("블로그(탭)", "블로그(탭)_URL", "블로그(탭)", "blog_rank_tab"),
]

PAGE_LOAD_TIMEOUT_MS_FIRST = 10000
PAGE_LOAD_TIMEOUT_MS_RETRY = 15000
PAGE_LOAD_RETRY_COUNT = 0
CURRENT_PAGE_LOAD_TIMEOUT_MS = PAGE_LOAD_TIMEOUT_MS_FIRST


def _set_page_load_timeout_for_attempt(attempt: int) -> int:
    global CURRENT_PAGE_LOAD_TIMEOUT_MS
    CURRENT_PAGE_LOAD_TIMEOUT_MS = (
        PAGE_LOAD_TIMEOUT_MS_FIRST if attempt == 0 else PAGE_LOAD_TIMEOUT_MS_RETRY
    )
    return CURRENT_PAGE_LOAD_TIMEOUT_MS


def normalize_blog_id(blog_id: str) -> str:
    """블로그 ID만 추출 (URL일 수 있음)."""
    s = blog_id.strip()
    m = re.search(r"blog\.naver\.com/([^/?]+)", s)
    if m:
        return m.group(1)
    return s.split("/")[-1] if "/" in s else s


def _normalize_blog_exposed_url(href: str, element=None) -> str | None:
    """
    카드 내 링크에서 실제 블로그 글 URL을 반환한다.
    - /p/crd/rd?...&u=https%3A%2F%2Fblog.naver.com%2Fjd_ah%2F224168368743 형태면 u 파라미터 디코딩
    - a 태그의 cru 속성에 실제 URL이 있으면 우선 사용 (네이버 카드에서 자주 사용)
    - 이미 https://blog.naver.com/... 형태면 그대로 사용 (쿼리 제거하여 글 주소만)
    """
    if not href or not href.strip():
        return None
    href = href.strip()
    # 1) cru 속성에 실제 노출 URL이 있는 경우 (예: cru="https://blog.naver.com/jd_ah/224168368743")
    if element:
        try:
            cru = element.get_attribute("cru")
            if cru and "blog.naver.com" in cru:
                return cru.strip()
        except Exception:
            pass
    # 2) 이미 직접 blog URL인 경우 (쿼리 제거)
    if "blog.naver.com" in href and (href.startswith("http://") or href.startswith("https://")):
        return href.split("?")[0].split("#")[0]
    # 3) 리다이렉트 URL에서 u= 파라미터로 실제 URL이 있는 경우
    if "u=" in href or "u%3D" in href or "u%3d" in href:
        try:
            parsed = urlparse(href)
            qs = parse_qs(parsed.query)
            for key in ("u", "U"):
                if key in qs and qs[key]:
                    raw = qs[key][0]
                    decoded = unquote(raw)
                    if "blog.naver.com" in decoded:
                        return decoded.split("?")[0].split("#")[0]
        except Exception:
            pass
    return href


# 글 주소 형태: https://blog.naver.com/아이디/글번호 (숫자로 끝)
_RE_BLOG_POST_URL = re.compile(r"https?://blog\.naver\.com/[^/]+/\d+", re.I)


def _get_best_blog_post_url_from_card(card, target_id: str) -> str | None:
    """
    카드 안에서 우리 블로그(target_id)로 가는 링크 중 '글 주소'(blog.naver.com/ID/숫자)를 우선 반환.
    반려동물 인기글·블로그 탭처럼 직접 href와 리다이렉트가 섞여 있어도 구체적인 URL을 골라준다.
    """
    candidates: list[str] = []
    for a in card.query_selector_all('a[href*="blog.naver.com"], a[cru*="blog.naver.com"], a[href*="/p/crd/rd"], a[href*="u="]'):
        href = (a.get_attribute("href") or "").strip()
        cru = (a.get_attribute("cru") or "").strip()
        url = None
        if cru and "blog.naver.com" in cru and target_id in cru:
            url = cru.split("?")[0].split("#")[0]
        if not url and href:
            url = _normalize_blog_exposed_url(href, a)
        if url and target_id in url and url not in candidates:
            candidates.append(url)
    # 글 주소 형태(blog.naver.com/ID/숫자)가 있으면 그걸 우선
    for u in candidates:
        if _RE_BLOG_POST_URL.search(u):
            return u
    return candidates[0] if candidates else None


def count_cards(page, container_selector: str, card_selector: str) -> int:
    """컨테이너 내의 카드 개수를 반환한다."""
    try:
        container = page.query_selector(container_selector)
        if not container:
            return 0
        cards = container.query_selector_all(card_selector)
        return len(cards)
    except Exception:
        return 0


def find_rank_in_cards(
    page, container_selector: str, card_selector: str, target_id: str, max_cards: int | None = None
) -> tuple[int | None, str | None]:
    """
    결과 '카드'(항목) 단위로 순위를 센다.
    - container 안에서 card_selector에 맞는 모든 카드를 DOM 순서대로 1, 2, 3...
    - 우리 블로그(blog.naver.com + target_id) 링크가 들어 있는 첫 번째 카드의 (순위, 노출 URL) 반환.
    - max_cards가 지정되면 해당 개수까지만 확인한다.
    """
    try:
        container = page.query_selector(container_selector)
        if not container:
            return None, None

        cards = container.query_selector_all(card_selector)
        if not cards:
            cards = container.query_selector_all('.fds-web-doc-root, [data-template-id="webItem"]')
        if max_cards is not None:
            cards = cards[:max_cards]

        for rank, card in enumerate(cards, 1):
            exposed_url: str | None = None

            links = card.query_selector_all('a[href*="blog.naver.com"]')
            for link in links:
                href = link.get_attribute("href") or ""
                if not href:
                    continue
                if "blog.naver.com" in href:
                    blog_id_match = re.search(r'blog\.naver\.com/([^/?]+)', href)
                    if blog_id_match and blog_id_match.group(1) == target_id:
                        return rank, _get_best_blog_post_url_from_card(card, target_id) or _normalize_blog_exposed_url(href, link) or href
                    if target_id in href:
                        return rank, _get_best_blog_post_url_from_card(card, target_id) or _normalize_blog_exposed_url(href, link) or href

            # 반려동물 인기글/블로그 탭: href에 blog가 인코딩되어 있거나 cru에만 있는 경우
            links_cru = card.query_selector_all('a[cru*="blog.naver.com"]')
            for link in links_cru:
                cru = link.get_attribute("cru") or ""
                if cru and target_id in cru and "blog.naver.com" in cru:
                    return rank, _get_best_blog_post_url_from_card(card, target_id) or cru.strip().split("?")[0].split("#")[0]

            links_rd = card.query_selector_all('a[href*="/p/crd/rd"], a[href*="u="]')
            for link in links_rd:
                href = link.get_attribute("href") or ""
                if not href:
                    continue
                normalized = _normalize_blog_exposed_url(href, link)
                if normalized and target_id in normalized:
                    return rank, _get_best_blog_post_url_from_card(card, target_id) or normalized

            links_in = card.query_selector_all('a[href*="in.naver.com"]')
            for link in links_in:
                href = link.get_attribute("href") or ""
                if not href:
                    continue
                in_match = re.search(r'in\.naver\.com/([^/?]+)', href)
                if in_match and in_match.group(1) == target_id:
                    return rank, _get_best_blog_post_url_from_card(card, target_id) or href

            card_html = card.inner_html() or ""
            if f"blog.naver.com/{target_id}" in card_html or f"in.naver.com/{target_id}" in card_html or target_id in card_html:
                best = _get_best_blog_post_url_from_card(card, target_id)
                if best:
                    return rank, best
                exposed_url = None
                first_blog = card.query_selector('a[href*="blog.naver.com"]')
                if first_blog:
                    h = first_blog.get_attribute("href")
                    exposed_url = _normalize_blog_exposed_url(h, first_blog) if h else None
                if not exposed_url:
                    first_cru = card.query_selector('a[cru*="blog.naver.com"]')
                    if first_cru:
                        cru = first_cru.get_attribute("cru")
                        if cru and target_id in cru:
                            exposed_url = cru.strip().split("?")[0].split("#")[0]
                if not exposed_url:
                    first_rd = card.query_selector('a[href*="/p/crd/rd"], a[href*="u="]')
                    if first_rd:
                        exposed_url = _normalize_blog_exposed_url(first_rd.get_attribute("href"), first_rd)
                if not exposed_url:
                    first_in = card.query_selector('a[href*="in.naver.com"]')
                    if first_in:
                        exposed_url = first_in.get_attribute("href")
                return rank, exposed_url or None

        return None, None
    except Exception as e:
        print(f"find_rank_in_cards 오류: {e}")
        return None, None


def _blog_tab_rank_fallback(page, target_id: str) -> tuple[int | None, str | None]:
    """블로그 탭 페이지에서 카드 구조를 못 찾을 때, 블로그 링크 순서로 순위·노출 URL 반환. 최대 20개까지만 확인. 글 주소 형태 우선."""
    try:
        container = page.query_selector("body")
        if not container:
            return None, None
        # 우리 블로그 링크만 수집 (직접 href + cru + 리다이렉트)
        candidates: list[tuple[int, str]] = []
        seen_urls: set[str] = set()
        for a in container.query_selector_all('a[href*="blog.naver.com"], a[cru*="blog.naver.com"], a[href*="/p/crd/rd"], a[href*="u="]'):
            href = a.get_attribute("href") or ""
            cru = (a.get_attribute("cru") or "").strip()
            url = None
            if cru and "blog.naver.com" in cru and target_id in cru:
                url = cru.split("?")[0].split("#")[0]
            if not url and href:
                url = _normalize_blog_exposed_url(href, a)
            if not url or target_id not in url or url in seen_urls:
                continue
            seen_urls.add(url)
            # 순위 = 지금까지 우리 블로그 URL 개수 (1부터)
            rank = len(candidates) + 1
            if rank > 20:
                break
            candidates.append((rank, url))
        if not candidates:
            return None, None
        # 글 주소 형태(blog.naver.com/ID/숫자)가 있으면 그걸 우선
        for r, u in candidates:
            if _RE_BLOG_POST_URL.search(u):
                return r, u
        return candidates[0][0], candidates[0][1]
    except Exception:
        return None, None


# --- 세 곳의 컨테이너 + 카드(결과 항목) 셀렉터 ---
# 순위 = "몇 번째 카드" (블로그/카페/인플루언서 등 섞여 있어도 카드 순서로 셈)

# 1) 통합검색 → 검색결과 섹션 (한 카드 = 한 결과: 블로그 or 카페 or 인플루언서 등)
# 진짜 "검색결과" 블록은 data-meta-area="web_gen", 노출 시에만 data-slog-visible="true"
SELECTOR_SEARCH_RESULT = [
    '#main_pack div[data-meta-area="web_gen"]',
]
CARD_SEARCH_RESULT = '.fds-web-doc-root'  # 검색결과 한 건 = 한 카드

# 2) 통합검색 → 반려동물 인기글 섹션 (한 카드 = 한 인기글 항목)
# 진짜 "반려동물 인기글"은 data-meta-area="ugB_bsR". kwL_ssT는 "함께 보면 좋은" 섹션이라 제외.
SELECTOR_PET_POPULAR = [
    '#main_pack [data-meta-area="ugB_bsR"]',
    '#main_pack .fds-ugc-single-intention-item-list',
]
# 더보기 클릭 후 열리는 레이어 안의 리스트 컨테이너 (스크롤 시 여기에 항목이 추가됨)
SELECTOR_PET_POPULAR_LAYER = [
    '.bridge_content._lb_open_target [data-meta-area="ugB_bsR"]',
    '.pack_group._lb_content_root [data-meta-area="ugB_bsR"]',
    '.bridge_content._lb_open_target .fds-ugc-single-intention-item-list',
]
CARD_PET_POPULAR = '[data-template-id="ugcItem"]'  # 인기글 한 건 = 한 카드

# 3) 블로그 탭 버튼 + 블로그 탭 페이지에서의 카드
SELECTOR_BLOG_TAB = 'a.tab[href*="tab.blog.all"]'
# 블로그 탭 페이지: 한 카드 = 한 블로그 결과 (직접 href 예: href="https://blog.naver.com/jd_ah/224143308359")
# 새 구조: section.sp_nblog 안에 [data-template-id="ugcItem"] 또는 ul.lst_view > ugcItem
SELECTOR_BLOG_TAB_CONTAINER = [
    '#main_pack section.sp_nblog',  # sds-comps 등 새 DOM에서 카드가 section 직하위에 있을 수 있음
    '#main_pack section.sp_nblog .api_subject_bx ul.lst_view',
    '#main_pack section.sp_nblog ul.lst_view',
    '#main_pack ul.lst_view',
]
CARD_BLOG_TAB = '[data-template-id="ugcItem"]'  # 블로그 탭 페이지의 각 블로그 결과 카드

# 4) 통합검색 → 일반 검색(통합 재정렬 rrB_bdR): 웹·리뷰·영상 혼합, 1/2/3... 페이지네이션
SELECTOR_GENERAL_SEARCH = [
    '#main_pack .spw_rerank[data-collection="rrB_bdR"]',
    '#main_pack [data-slog-container="rrB_bdR"]',
]
# 한 항목 = 한 블록 (data-fender-root + data-meta-area="rrB_bdR")
CARD_GENERAL_SEARCH = 'div[data-fender-root="true"][data-meta-area="rrB_bdR"]'

# --- 플레이스(스마트플레이스) 섹션: 광고 vs 일반 구분 ---
# 일반 예시: data-nmb_vcl-doc-id="31992363", 링크는 map.naver.com, 광고 라벨 없음
# 광고 예시: data-nmb_vcle-doc-id="1331794732_nad-a001-06-...", 링크는 ader.naver.com, "광고" 라벨 있음
#
# | 구분              | 일반 스마트플레이스        | 광고                          |
# |-------------------|----------------------------|-------------------------------|
# | doc-id 속성 이름  | data-nmb_vcl-doc-id        | data-nmb_vcle-doc-id (e 있음) |
# | doc-id 값         | 숫자만 (예: 31992363)     | _nad- 포함 (예: 1331794732_nad-a001-06-...) |
# | 링크 도메인       | map.naver.com              | ader.naver.com                |
# | 광고 라벨         | 없음                       | span.place_blind "광고", SVG place_ad_label_* |


def is_place_card_ad(card) -> bool:
    """
    플레이스 리스트의 한 항목(li)이 광고인지 여부를 판별한다.
    - 일반: data-nmb_vcl-doc-id (숫자만), map.naver.com 링크, 광고 라벨 없음
    - 광고: data-nmb_vcle-doc-id 또는 doc-id 값에 _nad- 포함, ader.naver.com 링크, "광고" 라벨
    card: Playwright ElementHandle (ul.zPw6U > li.c1sly 등)
    """
    try:
        # 1) 속성 이름이 nmb_vcle-doc-id 이면 광고 (일반은 nmb_vcl-doc-id)
        if card.get_attribute("data-nmb_vcle-doc-id"):
            return True
        # 2) data-nmb_vcl-doc-id 값에 _nad- 가 포함되면 광고
        doc_id = card.get_attribute("data-nmb_vcl-doc-id") or ""
        if "_nad-" in doc_id:
            return True
        # 3) 카드 내부에 ader.naver.com 링크가 있으면 광고 (일반은 map.naver.com만 사용)
        if card.query_selector('a[href*="ader.naver.com"]'):
            return True
        # 4) "광고" 라벨 요소가 있으면 광고 (place_ad_label SVG 또는 place_blind "광고")
        if card.query_selector(".place_ad_label_border, .place_ad_label_text"):
            return True
        blind = card.query_selector('span.place_blind')
        if blind and (blind.inner_text() or "").strip() == "광고":
            return True
        return False
    except Exception:
        return False


# 플레이스 섹션 두 가지 레이아웃:
#  (A) 표준 팩: #place-main-section-root → ul.zPw6U → li.c1sly, 이름 span.jVsoy
#  (B) 지역형:  #loc-main-section-root → div.wObwH 카드, 이름 span.Ypcqn (넓은 지역/진료 키워드에서 이 레이아웃으로 그려짐)
# 광고 제외 후 순위 카운트.
SELECTOR_PLACE_CONTAINER = [
    "#place-main-section-root ul.zPw6U",
    "#place-main-section-root ul[class*='zPw6U']",
    "ul.zPw6U",
    "#loc-main-section-root",  # 지역형(B) — 없으면 다음으로, 있으면 여기서 카드 탐색
]
# 한 항목: (A) li.c1sly / data-nmb_vcl-doc-id, (B) div.wObwH
CARD_PLACE = "li.c1sly, div.wObwH"


def _normalize_store_name(s: str) -> str:
    """상호명 비교용: 공백 정리, &/&amp; 통일."""
    if not s:
        return ""
    s = (s or "").strip().replace("&amp;", "&").replace("&", " ")
    return " ".join(s.split())


def _place_debug_on() -> bool:
    """RANK_PLACE_DEBUG 가 켜져 있으면 플레이스 파싱 진단을 stderr 로 출력."""
    v = (os.getenv("RANK_PLACE_DEBUG") or "").strip().lower()
    return v in ("1", "true", "yes", "on")


def _place_card_name(li) -> str:
    """
    카드(li.c1sly / div.wObwH ...)에서 상호명을 뽑는다.
    네이버가 난독화 클래스(span.jVsoy·span.Ypcqn)를 자주 바꾸므로, 못 찾으면
    카드 전체 텍스트로 폴백한다. 매칭은 부분일치라 카드 텍스트 안에 상호명이 있으면 잡힌다.
    """
    el = li.query_selector("span.jVsoy, span.Ypcqn")
    if el:
        t = (el.inner_text() or "").strip()
        if t:
            return t
    # 폴백: 이름 span 을 못 찾음(클래스 변경 등) → 카드 전체 텍스트.
    try:
        return (li.inner_text() or "").strip()
    except Exception:
        return ""


def _store_name_matches(target_norm: str, card_name: str) -> bool:
    """
    타깃(정규화된 상호명)이 카드 이름/텍스트와 일치하는지.
    - 정규화 후 부분일치(양방향)
    - 공백 위치 차이 무시(네이버 표기 vs DB 등록명): '허브 동물메디컬센터' ↔ '허브동물메디컬센터'
    짧은 문자열 오탐을 줄이려 역방향(카드명 ⊂ 타깃)은 4자 이상일 때만 허용.
    """
    if not target_norm or not card_name:
        return False
    nn = _normalize_store_name(card_name)
    if not nn:
        return False
    if target_norm == nn or target_norm in nn or target_norm in card_name:
        return True
    td = "".join(target_norm.split())
    nd = "".join(nn.split())
    if not td or not nd:
        return False
    if td == nd or td in nd:
        return True
    if len(nd) >= 4 and nd in td:
        return True
    return False


def _expand_place_section(page, container_sel) -> None:
    """
    지역형(loc) 플레이스 섹션은 초기 ~5개만 HTML에 그리고 나머지는 스크롤/'더보기' 로 지연 로딩한다.
    카드 수가 더 안 늘 때까지 (1) 마지막 카드로 스크롤 (2) '더보기' 클릭 을 반복해 전부 로드시킨다.
    표준 팩(zPw6U)에는 호출하지 않으므로 기존 페이지네이션 동작에는 영향 없음.
    """
    try:
        for _ in range(10):
            cont = page.query_selector(container_sel)
            if not cont:
                return
            cards = cont.query_selector_all(CARD_PLACE)
            before = len(cards)
            # (1) 마지막 카드/섹션 끝으로 스크롤 → 지연 로딩 트리거
            if cards:
                try:
                    cards[-1].scroll_into_view_if_needed(timeout=1500)
                except Exception:
                    pass
            try:
                page.evaluate("window.scrollBy(0, 1200)")
            except Exception:
                pass
            page.wait_for_timeout(500)
            # (2) '더보기' 요소가 있으면 클릭(클래스 비의존 — 텍스트로 탐색)
            cont = page.query_selector(container_sel)
            if cont:
                for el in cont.query_selector_all("a, button"):
                    tt = (el.inner_text() or "").strip()
                    if tt and "더보기" in tt:
                        try:
                            el.click(timeout=1500)
                            page.wait_for_timeout(700)
                        except Exception:
                            pass
                        break
            cont = page.query_selector(container_sel)
            after = len(cont.query_selector_all(CARD_PLACE)) if cont else before
            if _place_debug_on() and after != before:
                print(f"   [place-debug] 섹션 확장: 카드 {before} → {after}", file=sys.stderr)
            if after <= before:
                return  # 더 안 늘어남 → 전부 로드됨
    except Exception:
        return


def find_place_rank_in_page(page, container_selector: str, store_name: str) -> tuple[int | None, str | None]:
    """
    플레이스 리스트가 있는 페이지에서 '상호명'과 일치하는 항목의 순위(1부터)와 노출 URL을 반환.
    광고(li)는 순위에서 제외하고, 일반 항목만 1, 2, 3... 카운트.
    """
    try:
        ul = page.query_selector(container_selector)
        if not ul:
            return None, None
        items = ul.query_selector_all(CARD_PLACE)
        if not items:
            items = ul.query_selector_all("li[data-nmb_vcl-doc-id], li[data-nmb_vcle-doc-id]")
        target = _normalize_store_name(store_name)
        rank = 0
        for li in items:
            if is_place_card_ad(li):
                continue
            rank += 1
            name = _place_card_name(li)
            if _place_debug_on():
                print(f"   [place-debug] #{rank} name={name[:60]!r}", file=sys.stderr)
            if _store_name_matches(target, name):
                return rank, None  # 플레이스는 URL 수집하지 않음
        return None, None
    except Exception:
        return None, None


def get_place_rank_with_pagination(page, keyword: str, store_name: str, integrated_url: str) -> tuple[int | None, str | None]:
    """
    통합검색 페이지에서 플레이스 섹션을 찾고, 상호명이 일치할 때까지
    현재 페이지 → 다음 페이지 순으로 순위를 찾는다. 광고 제외.
    """
    try:
        container_sel = None
        for sel in SELECTOR_PLACE_CONTAINER:
            if page.query_selector(sel):
                container_sel = sel
                break
        if not container_sel:
            return None, None

        total_rank_offset = 0  # 이전 페이지까지의 (광고 제외) 개수

        while True:
            rank_on_page, url = find_place_rank_in_page(page, container_sel, store_name)
            if rank_on_page is not None:
                return total_rank_offset + rank_on_page, url

            # 현재 페이지의 광고 제외 일반 항목 개수만큼 오프셋 증가
            ul = page.query_selector(container_sel)
            if not ul:
                break
            items = ul.query_selector_all(CARD_PLACE) or ul.query_selector_all("li[data-nmb_vcl-doc-id], li[data-nmb_vcle-doc-id]")
            for li in items:
                if not is_place_card_ad(li):
                    total_rank_offset += 1

            next_btn = page.query_selector('div.cmm_pgs.x5Efp a.cmm_pg_next:not([aria-disabled="true"])')
            if not next_btn:
                break
            next_btn.click()
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            page.wait_for_timeout(300)
            container_sel = None
            for sel in SELECTOR_PLACE_CONTAINER:
                if page.query_selector(sel):
                    container_sel = sel
                    break
            if not container_sel:
                break
        return None, None
    except Exception:
        return None, None


def _place_first_card_text(page, container_sel: str) -> str:
    """플레이스 목록 첫 카드의 상호명(없으면 앞부분 텍스트). AJAX 페이지 전환 감지용 시그니처."""
    try:
        ul = page.query_selector(container_sel)
        if not ul:
            return ""
        cards = ul.query_selector_all(CARD_PLACE)
        if not cards:
            cards = ul.query_selector_all("li[data-nmb_vcl-doc-id], li[data-nmb_vcle-doc-id]")
        if not cards:
            return ""
        el = cards[0].query_selector("span.jVsoy, span.Ypcqn")
        if el:
            return (el.inner_text() or "").strip()
        return (cards[0].inner_text() or "").strip()[:50]
    except Exception:
        return ""


def get_place_ranks_with_pagination(page, keyword: str, targets: list[str]) -> dict[str, int | None]:
    """통합검색 플레이스 섹션에서 '여러 상호명'의 순위를 한 번의 탐색(페이지네이션)으로 찾는다.
    우리 병원 + 경쟁사를 한 검색에서 같이 찾기 위함. 반환: {원본 target 문자열: rank or None}. 광고 제외."""
    result: dict[str, int | None] = {t: None for t in targets if t}
    norm_map: dict[str, str] = {}  # normalized store name -> original
    for t in targets:
        if t:
            norm_map[_normalize_store_name(t)] = t
    if not norm_map:
        return result
    try:
        container_sel = None
        for sel in SELECTOR_PLACE_CONTAINER:
            if page.query_selector(sel):
                container_sel = sel
                break
        if not container_sel:
            if _place_debug_on():
                print(
                    f"   [place-debug] '{keyword}': 플레이스 컨테이너 없음 "
                    f"(시도한 셀렉터: {SELECTOR_PLACE_CONTAINER}) — 레이아웃이 또 바뀌었을 수 있음",
                    file=sys.stderr,
                )
            return result

        total_rank_offset = 0
        # 우리 가게가 깊은 순위(여러 페이지 뒤)일 수 있어 넉넉히 넘긴다. (원래 단일타깃은 무제한이었음)
        # 우리 가게를 찾은 뒤에는 경쟁사용으로 몇 페이지만 더 보고 멈춰 런타임을 제한한다.
        max_pages = int(os.getenv("PLACE_MAX_PAGES", "12") or "12")
        store_orig = targets[0] if targets else None  # 우리 병원(첫 타깃)
        # 우리 가게를 찾은 뒤 경쟁사를 위해 더 보는 페이지 수. 경쟁사는 보통 우리보다 하위라 충분히 봐야 함.
        extra_pages_after_store = int(os.getenv("PLACE_EXTRA_PAGES_AFTER_STORE", "6") or "6")
        pages_after_store = 0
        for _ in range(max_pages):
            # 지역형(loc) 레이아웃은 초기 ~5개만 그려지고 나머지는 지연 로딩 → 읽기 전에 전부 펼친다.
            # (표준 팩 zPw6U 는 페이지네이션으로 처리하므로 확장하지 않는다.)
            if container_sel == "#loc-main-section-root":
                _expand_place_section(page, container_sel)
            ul = page.query_selector(container_sel)
            if not ul:
                break
            items = ul.query_selector_all(CARD_PLACE)
            if not items:
                items = ul.query_selector_all("li[data-nmb_vcl-doc-id], li[data-nmb_vcle-doc-id]")
            rank = total_rank_offset
            for li in items:
                if is_place_card_ad(li):
                    continue
                rank += 1
                name = _place_card_name(li)
                if _place_debug_on():
                    print(f"   [place-debug] '{keyword}' #{rank} name={name[:60]!r}", file=sys.stderr)
                for nt, orig in norm_map.items():
                    if result.get(orig) is not None:
                        continue
                    if _store_name_matches(nt, name):
                        result[orig] = rank
            total_rank_offset = rank
            if all(result.get(o) is not None for o in norm_map.values()):
                break
            # 우리 가게를 이미 찾았으면 경쟁사 위해 몇 페이지만 더 보고 중단(런타임 절약)
            if store_orig is not None and result.get(store_orig) is not None:
                pages_after_store += 1
                if pages_after_store > extra_pages_after_store:
                    break
            next_btn = page.query_selector('div.cmm_pgs.x5Efp a.cmm_pg_next:not([aria-disabled="true"])')
            if not next_btn:
                # 표준 '다음페이지' 버튼이 없음 → 지역형(loc) 레이아웃일 수 있음. '더보기' 요소 후보를
                # 덤프해서(클래스·href·tag) 지역형 페이지네이션을 구현할 근거를 남긴다.
                if _place_debug_on():
                    cont = page.query_selector(container_sel)
                    cards_n = len(cont.query_selector_all(CARD_PLACE)) if cont else 0
                    cands = []
                    scope = cont or page
                    try:
                        for el in scope.query_selector_all("a, button"):
                            tt = (el.inner_text() or "").strip()
                            if tt and "더보기" in tt:
                                tag = el.evaluate("e => e.tagName")
                                cls = el.get_attribute("class") or ""
                                href = el.get_attribute("href") or ""
                                role = el.get_attribute("role") or ""
                                cands.append(f"<{tag} text={tt[:24]!r} class={cls[:60]!r} role={role!r} href={href[:70]!r}>")
                    except Exception as _e:
                        cands.append(f"(enum 실패: {_e})")
                    print(
                        f"   [place-debug] '{keyword}': 다음버튼 없음 → 멈춤. container={container_sel}, "
                        f"카드수={cards_n}, 더보기후보={cands if cands else '없음'}",
                        file=sys.stderr,
                    )
                break
            prev_sig = _place_first_card_text(page, container_sel)
            next_btn.click()
            # AJAX 페이지 전환: 고정 대기(과거 300ms) 대신 '첫 카드가 바뀔 때까지' 폴링.
            # 느린 환경/Chrome 경합에서도 다음 페이지를 확실히 읽어 깊은 순위 누락을 막는다.
            changed = False
            for _ in range(24):  # 최대 ~6초
                page.wait_for_timeout(250)
                new_sel = None
                for sel in SELECTOR_PLACE_CONTAINER:
                    if page.query_selector(sel):
                        new_sel = sel
                        break
                if not new_sel:
                    continue
                if _place_first_card_text(page, new_sel) != prev_sig:
                    container_sel = new_sel
                    changed = True
                    break
            if not changed:
                break  # 페이지 내용이 안 바뀜(전환 실패) → 더 진행 무의미
        return result
    except Exception:
        return result


def _is_truthy_env(name: str) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return False
    return str(raw).strip().lower() in {"1", "true", "yes", "y", "on"}


def create_browser_session(playwright, *, headless: bool = True, use_debug_chrome: bool = False, debug_port: int = 9222):
    """
    브라우저 세션 생성:
    - 기본: 독립 Chromium launch
    - CDP 모드: 디버깅 Chrome에 접속
    """
    if use_debug_chrome:
        endpoint = f"http://127.0.0.1:{debug_port}"
        # connect_over_cdp 의 첫 시도가 간헐적으로 실패하는 경우가 있다(크롬은 떠 있는데 웹소켓 연결 실패).
        # 실측상 첫 워커(블로그)는 실패하고 바로 뒤 워커(플레이스)는 성공하는 패턴 → 짧게 재시도하면 붙는다.
        browser = None
        last_err: Exception | None = None
        max_attempts = int(os.getenv("RANK_CDP_CONNECT_RETRIES", "4") or "4")
        for attempt in range(1, max_attempts + 1):
            try:
                browser = playwright.chromium.connect_over_cdp(endpoint)
                break
            except Exception as e:
                last_err = e
                if attempt < max_attempts:
                    print(f"⏳ CDP 연결 재시도 {attempt}/{max_attempts - 1} — {endpoint}", file=sys.stderr)
                    time.sleep(1.5 * attempt)  # 점증 대기(1.5s → 3s → 4.5s …)
        if browser is None:
            raise RuntimeError(
                f"디버깅 Chrome(CDP) 연결 실패: {endpoint} ({max_attempts}회 재시도). "
                f"Chrome을 --remote-debugging-port={debug_port}로 실행했는지 확인하세요."
            ) from last_err

        context = browser.contexts[0] if browser.contexts else browser.new_context()
        context.route("**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,otf,eot}", lambda route: route.abort())

        def cleanup():
            # CDP 모드에서는 기존 사용중인 Chrome(및 기존 탭)을 건드리지 않는다.
            # (contexts[0]가 사용자의 기본 컨텍스트일 수 있어 close 하면 탭이 닫힐 수 있음)
            with suppress(Exception):
                browser.close()

        return browser, context, cleanup

    # 자동화 신호(--enable-automation, navigator.webdriver) 제거로 봇 탐지 회피.
    browser = playwright.chromium.launch(
        headless=headless,
        timeout=30000,
        args=["--disable-blink-features=AutomationControlled"],
    )
    # UA 가 실제 설치 Chrome 버전과 어긋나면 핑거프린트 불일치 신호가 된다. 최신값을 기본으로 두되,
    # 워커 PC 의 실제 Chrome 버전에 맞추고 싶으면 RANK_USER_AGENT 로 override.
    ua = os.getenv("RANK_USER_AGENT") or (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )
    context = browser.new_context(user_agent=ua)
    # launch 모드에서 true 로 노출되는 navigator.webdriver 흔적 제거(CDP 모드는 원래 false).
    context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")
    context.route("**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,otf,eot}", lambda route: route.abort())

    def cleanup():
        with suppress(Exception):
            context.close()
        with suppress(Exception):
            browser.close()

    return browser, context, cleanup


def check_place_ranking(
    keyword: str,
    store_name: str,
    headless: bool = True,
    use_debug_chrome: bool = False,
    debug_port: int = 9222,
) -> dict:
    """
    키워드로 네이버 검색 후 플레이스 섹션에서 상호명이 일치하는 항목의 순위를 찾는다.
    광고는 순위에서 제외하며, 2·3페이지까지 페이지네이션한다.
    """
    encoded_kw = quote(keyword)
    integrated_url = f"https://search.naver.com/search.naver?query={encoded_kw}"
    out = {"keyword": keyword, "store_name": store_name, "rank": None, "url": None, "error": None}

    with sync_playwright() as p:
        browser, context, cleanup = create_browser_session(
            p,
            headless=headless,
            use_debug_chrome=use_debug_chrome,
            debug_port=debug_port,
        )
        try:
            page = context.new_page()
            out = check_place_ranking_on_page(page, keyword, store_name)
        finally:
            cleanup()
    return out


def check_place_ranking_on_page(page, keyword: str, store_name: str, competitor_stores: list[str] | None = None) -> dict:
    """
    주입된 Playwright page를 재사용하며 플레이스 순위를 확인한다.
    (탭 누적 방지를 위해 page를 새로 만들지 않는다.)
    competitor_stores 가 주어지면 같은 한 번의 검색에서 경쟁사 상호 순위도 함께 찾는다.
    """
    competitor_stores = [s for s in (competitor_stores or []) if s]
    encoded_kw = quote(keyword)
    integrated_url = f"https://search.naver.com/search.naver?query={encoded_kw}"
    out = {"keyword": keyword, "store_name": store_name, "rank": None, "url": None, "error": None, "competitor_ranks": {}}

    for attempt in range(PAGE_LOAD_RETRY_COUNT + 1):
        timeout_ms = _set_page_load_timeout_for_attempt(attempt)
        try:
            page.goto(integrated_url, wait_until="domcontentloaded", timeout=timeout_ms)
            page.wait_for_timeout(100)
            ranks = get_place_ranks_with_pagination(page, keyword, [store_name] + competitor_stores)
            out["rank"] = ranks.get(store_name)
            out["url"] = None
            out["competitor_ranks"] = {s: ranks.get(s) for s in competitor_stores}
            out["error"] = None
            break
        except PlaywrightTimeout:
            out["error"] = "페이지 로드 시간 초과"
            if attempt >= PAGE_LOAD_RETRY_COUNT:
                break
        except Exception as e:
            out["error"] = str(e)
            break
    return out


def try_click_more_and_find_rank_blog_tab(
    page, container_selector: str, card_selector: str, more_button_selector: str, target_id: str
) -> tuple[int | None, str | None]:
    """
    블로그 탭 "더보기" 버튼을 반복 클릭하여 추가 결과에서 순위를 찾는다.
    최대 20개 카드까지만 확인. 반환: (순위, 노출 URL).
    """
    try:
        max_cards = 20
        current_count = count_cards(page, container_selector, card_selector)
        if current_count >= max_cards:
            return find_rank_in_cards(page, container_selector, card_selector, target_id, max_cards)
        
        # "더보기" 버튼을 반복 클릭하면서 20개까지 확인
        while current_count < max_cards:
            more_button = page.query_selector(more_button_selector)
            if not more_button:
                # 더보기 버튼이 없으면 현재까지의 결과에서만 확인
                break
            
            # 버튼 클릭
            more_button.click()
            page.wait_for_timeout(100)  # 추가 결과 로드 대기
            
            # 카드 수 다시 확인
            new_count = count_cards(page, container_selector, card_selector)
            
            # 카드가 추가되지 않았으면 중단
            if new_count == current_count:
                break
            
            current_count = new_count
            
            if current_count >= max_cards:
                break
        return find_rank_in_cards(page, container_selector, card_selector, target_id, max_cards)
    except Exception:
        return None, None


def try_click_more_and_find_rank_search_result(
    page, more_button_selector: str, target_id: str
) -> int | None:
    """
    검색결과 "더보기" 버튼 클릭 후 페이지 이동하여 추가 결과에서 순위를 찾는다.
    검색결과 "더보기"는 페이지 이동이므로, 새 페이지에서 web_lis 컨테이너를 찾아야 함.
    2페이지의 맨 위 카드부터 1번으로 카운트한다.
    최대 20개 카드까지만 확인한다. 2페이지에 20개가 안 나오면 3페이지로 넘어간다.
    """
    try:
        max_cards = 20
        
        # "더보기" 버튼 찾기 — 여러 셀렉터 시도 (키워드별 DOM 구조 차이 대응)
        clicked = False
        more_button = page.query_selector(more_button_selector)
        if more_button:
            more_button.click()
            clicked = True
        if not clicked:
            for fallback_sel in [
                '#main_pack a[data-heatmap-target=".more2"]',
                'a[data-heatmap-target=".more2"]',
                '[data-meta-area="web_gen"] a[data-heatmap-target=".more2"]',
            ]:
                more_button = page.query_selector(fallback_sel)
                if more_button:
                    more_button.click()
                    clicked = True
                    break
        if not clicked:
            # 텍스트로 찾기 (검색결과 더보기)
            loc = page.locator('a:has-text("검색결과 더보기")')
            if loc.count() > 0:
                loc.first.click()
                clicked = True
        if not clicked:
            return None, None

        # 2페이지 로드 대기
            page.wait_for_load_state("domcontentloaded", timeout=CURRENT_PAGE_LOAD_TIMEOUT_MS)
        page.wait_for_timeout(100)  # 추가 결과 로드 대기

        # 새 페이지에서 web_lis 컨테이너 찾기 (검색결과 전용 페이지, 2페이지)
        # 여러 셀렉터 시도
        container_sel = None
        for sel in [
            '#main_pack div[data-meta-area="web_lis"]',
            '#main_pack [data-meta-area="web_lis"]',
            'div[data-meta-area="web_lis"]',
        ]:
            container = page.query_selector(sel)
            if container:
                container_sel = sel
                break
        
        if not container_sel:
            return None, None

        # 컨테이너가 나타날 때까지 대기
        try:
            page.wait_for_selector(container_sel, timeout=5000)
        except:
            pass
        
        # 카드들이 로드될 때까지 추가 대기
        page.wait_for_timeout(100)
        
        # 카드가 실제로 로드되었는지 확인
        cards = page.query_selector_all(f'{container_sel} {CARD_SEARCH_RESULT}')
        if not cards:
            # 대체 셀렉터로 카드 찾기
            cards = page.query_selector_all(f'{container_sel} .fds-web-doc-root')
        if not cards:
            # 추가 대기 후 다시 시도
            page.wait_for_timeout(100)
            cards = page.query_selector_all(f'{container_sel} {CARD_SEARCH_RESULT}')
        
        rank, url = find_rank_in_cards(page, container_sel, CARD_SEARCH_RESULT, target_id, max_cards)
        if rank is not None:
            return rank, url

        page2_count = count_cards(page, container_sel, CARD_SEARCH_RESULT)
        remaining_cards = max_cards - page2_count
        if page2_count < max_cards and remaining_cards > 0:
            page3_button = page.query_selector('a[href*="page=3"], .paging a:has-text("3"), .paging a.next')
            if page3_button:
                page3_button.click()
                page.wait_for_load_state("domcontentloaded", timeout=CURRENT_PAGE_LOAD_TIMEOUT_MS)
                page.wait_for_timeout(100)
                rank_on_page3, url_on_page3 = find_rank_in_cards(
                    page, container_sel, CARD_SEARCH_RESULT, target_id, remaining_cards
                )
                if rank_on_page3 is not None:
                    return page2_count + rank_on_page3, url_on_page3
        return None, None
    except Exception:
        return None, None


def try_click_more_and_find_rank_pet_popular(
    page, container_selector: str, card_selector: str, more_button_selector: str, target_id: str
) -> tuple[int | None, str | None]:
    """
    반려동물 인기글 "더보기" 클릭 후, 레이어가 열리면 레이어 내 리스트를 기준으로
    스크롤(마우스 휠)로 추가 로드되는 카드까지 확인한다. 최대 20개 카드까지만 확인.
    """
    try:
        max_cards = 20

        current_count = count_cards(page, container_selector, card_selector)
        if current_count >= max_cards:
            return find_rank_in_cards(page, container_selector, card_selector, target_id, max_cards)

        # 1) "더보기" 버튼 클릭 → 레이어(모달) 열림
        more_button = page.query_selector(more_button_selector)
        if more_button:
            more_button.click()
            page.wait_for_timeout(100)

        # 2) 레이어가 열렸다면 레이어 안의 리스트 컨테이너를 사용 (스크롤 시 여기에 항목 추가됨)
        effective_selector = container_selector
        layer_container = None
        for layer_sel in SELECTOR_PET_POPULAR_LAYER:
            el = page.query_selector(layer_sel)
            if el:
                effective_selector = layer_sel
                layer_container = el
                break
        if layer_container is None:
            layer_container = page.query_selector(container_selector)

        current_count = count_cards(page, effective_selector, card_selector)

        # 3) 스크롤 반복 — 레이어는 마우스 휠/스크롤에 반응해 추가 로드
        if layer_container:
            no_change_count = 0
            scroll_round = 0
            while current_count < max_cards and no_change_count < 4 and scroll_round < 15:
                prev_count = current_count
                scroll_round += 1
                try:
                    box = layer_container.bounding_box()
                    if box:
                        # 레이어 영역 중앙에서 마우스 휠로 스크롤 (무한스크롤 트리거)
                        cx = box["x"] + box["width"] / 2
                        cy = box["y"] + min(200, box["height"] / 2)
                        page.mouse.move(cx, cy)
                        for _ in range(3):
                            page.mouse.wheel(0, 400)
                            page.wait_for_timeout(75)
                    page.evaluate("window.scrollBy(0, 350)")
                    page.wait_for_timeout(50)
                    layer_container.evaluate("el => { el.scrollTop = el.scrollHeight; }")
                    page.wait_for_timeout(50)
                except Exception:
                    pass
                current_count = count_cards(page, effective_selector, card_selector)
                if current_count > prev_count:
                    no_change_count = 0
                else:
                    no_change_count += 1

        return find_rank_in_cards(page, effective_selector, card_selector, target_id, max_cards)
    except Exception:
        return None, None


def try_find_rank_general_search(
    page, target_id: str, integrated_url: str
) -> tuple[int | None, bool, str | None, bool]:
    """
    일반 검색(rrB_bdR) 섹션에서 순위를 찾는다. 1페이지에서 먼저 확인하고,
    없으면 2페이지로 넘어가며 최대 20개 항목까지만 확인한다.
    반환: (순위 또는 None, 섹션 존재 여부, 노출 URL 또는 None, 페이지 이동 여부)
    """
    max_cards = 20
    container_sel = None
    for sel in SELECTOR_GENERAL_SEARCH:
        if page.query_selector(sel):
            container_sel = sel
            break
    if not container_sel:
        return None, False, None, False

    try:
        rank, url = find_rank_in_cards(
            page, container_sel, CARD_GENERAL_SEARCH, target_id, max_cards=max_cards
        )
        if rank is not None:
            return rank, True, url, False  # 1페이지에서 발견 — 이동 없음

        count = count_cards(page, container_sel, CARD_GENERAL_SEARCH)
        if count >= max_cards:
            return None, True, None, False

        # 2페이지로 이동 시도 (start=11 또는 페이지 번호 2 링크)
        next_link = page.query_selector(
            'a[href*="start=11"], .spw_rerank a[href*="start="], '
            '.paging a:has-text("2"), a.pg_next, [data-collection="rrB_bdR"] a[href*="start="]'
        )
        if not next_link and page.locator('a:has-text("2")').count() > 0:
            next_link = page.locator('a:has-text("2")').first
        if not next_link:
            return None, True, None, False
        try:
            next_link.click()
        except Exception:
            return None, True, None, False

        page.wait_for_load_state("domcontentloaded", timeout=CURRENT_PAGE_LOAD_TIMEOUT_MS)
        page.wait_for_timeout(100)
        container_sel = None
        for sel in SELECTOR_GENERAL_SEARCH:
            if page.query_selector(sel):
                container_sel = sel
                break
        if not container_sel:
            return None, True, None, True  # 2페이지로 이동했지만 섹션 없음

        remaining = max_cards - count
        rank_on_page, url_on_page = find_rank_in_cards(
            page, container_sel, CARD_GENERAL_SEARCH, target_id, max_cards=remaining
        )
        if rank_on_page is not None:
            return count + rank_on_page, True, url_on_page, True
        return None, True, None, True
    except Exception:
        return None, True, None, False


def _close_pet_popular_layer_if_open(page) -> None:
    """
    반려동물 인기글 "더보기"로 연 레이어가 열려 있으면 닫는다.
    일반 검색(rrB_bdR)은 배경(통합검색 본문)에 있으므로, 레이어를 닫아야 탐색 가능.
    """
    try:
        # 1) Escape 키로 레이어 닫기 시도 (대부분의 모달/레이어 공통)
        page.keyboard.press("Escape")
        page.wait_for_timeout(50)
        # 2) 레이어 전용 닫기 버튼 클릭 시도 (네이버 LayerBridge 등)
        for close_sel in [
            'button[aria-label="닫기"]',
            '[class*="close"]',
            '._lb_close',
            '.layer_close',
            '.bridge_content ~ button',
            '.spw_rerank .btn_close',
            'a[role="button"]:has-text("닫기")',
        ]:
            btn = page.query_selector(close_sel)
            if btn:
                try:
                    btn.click()
                    page.wait_for_timeout(50)
                    break
                except Exception:
                    pass
        # 3) 백드롭(딤) 클릭으로 레이어 닫기 시도
        backdrop = page.query_selector('.bridge_backdrop, .layer_backdrop, [class*="backdrop"]')
        if backdrop:
            try:
                backdrop.click()
                page.wait_for_timeout(50)
            except Exception:
                pass
    except Exception:
        pass


def get_three_section_ranks(page, target_id: str, integrated_url: str) -> dict[str, int | None]:
    """
    통합검색(전체 탭) 페이지에서 순서대로:
    1) 검색결과: 통합검색 내 순위 먼저 확인 → 없으면 더보기 클릭 후 2페이지에서 20등까지 확인 → 1페이지로 복귀
    2) 반려동물 인기글: 통합검색 내 순위 먼저 확인 → 없으면 더보기 클릭 후 20등까지 확인 → (레이어 열렸으면) 레이어 닫기
    3) 일반 검색(rrB_bdR): 1페이지에서 확인 → 없으면 2페이지까지 최대 20개 확인
    블로그(탭)은 check_naver_ranking에서 별도 처리.
    """
    result: dict[str, int | None] = {}

    # --- 1) 검색결과: 통합검색 내 순위 먼저 → 없으면 더보기 후 2페이지에서 20등까지 ---
    _t0 = time.perf_counter()
    result["검색결과"] = None
    result["검색결과_URL"] = None
    search_result_exists = False
    page.wait_for_timeout(100)
    try:
        page.wait_for_selector('#main_pack a[data-heatmap-target=".more2"]', state="visible", timeout=1500)
    except Exception:
        try:
            page.locator('a:has-text("검색결과 더보기")').first.wait_for(state="visible", timeout=1500)
        except Exception:
            pass
    search_more_clicked = False
    for sel in SELECTOR_SEARCH_RESULT:
        container = page.query_selector(sel)
        if container:
            search_result_exists = True
            result["검색결과"], result["검색결과_URL"] = find_rank_in_cards(
                page, sel, CARD_SEARCH_RESULT, target_id
            )
            if result["검색결과"] is None:
                more_button_sel = f'{sel} a[data-heatmap-target=".more2"]'
                result["검색결과"], result["검색결과_URL"] = try_click_more_and_find_rank_search_result(
                    page, more_button_sel, target_id
                )
                search_more_clicked = True
            break
    if not search_result_exists:
        if page.query_selector('a[data-heatmap-target=".more2"]') or page.locator('a:has-text("검색결과 더보기")').count() > 0:
            search_result_exists = True
            result["검색결과"], result["검색결과_URL"] = find_rank_in_cards(
                page, SELECTOR_SEARCH_RESULT[0], CARD_SEARCH_RESULT, target_id
            )
            if result["검색결과"] is None:
                result["검색결과"], result["검색결과_URL"] = try_click_more_and_find_rank_search_result(
                    page, 'a[data-heatmap-target=".more2"]', target_id
                )
                search_more_clicked = True
    result["_검색결과_섹션있음"] = search_result_exists

    # 더보기를 실제로 눌러서 페이지를 이동했을 때만 통합검색으로 복귀
    if search_more_clicked:
        page.goto(integrated_url, wait_until="domcontentloaded", timeout=CURRENT_PAGE_LOAD_TIMEOUT_MS)
        page.wait_for_timeout(100)
    result["_t_검색결과"] = time.perf_counter() - _t0

    # --- 2) 반려동물 인기글 ---
    _t0 = time.perf_counter()
    result["반려동물 인기글"] = None
    result["반려동물 인기글_URL"] = None
    pet_exists = False
    for sel in SELECTOR_PET_POPULAR:
        container = page.query_selector(sel)
        if container:
            pet_exists = True
            result["반려동물 인기글"], result["반려동물 인기글_URL"] = find_rank_in_cards(
                page, sel, CARD_PET_POPULAR, target_id
            )
            if result["반려동물 인기글"] is None:
                more_button_sel = f'{sel} a[data-heatmap-target=".more"]'
                result["반려동물 인기글"], result["반려동물 인기글_URL"] = try_click_more_and_find_rank_pet_popular(
                    page, sel, CARD_PET_POPULAR, more_button_sel, target_id
                )
            break
    result["_반려동물_섹션있음"] = pet_exists

    # 반려동물 인기글 "더보기" 레이어가 열려 있으면 닫기 → 일반 검색은 배경 페이지에서 탐색
    _close_pet_popular_layer_if_open(page)
    result["_t_반려동물 인기글"] = time.perf_counter() - _t0

    # --- 3) 일반 검색(rrB_bdR) ---
    _t0 = time.perf_counter()
    result["일반 검색"] = None
    result["일반 검색_URL"] = None
    general_rank, general_exists, general_url, general_navigated = try_find_rank_general_search(page, target_id, integrated_url)
    result["일반 검색"] = general_rank
    result["일반 검색_URL"] = general_url
    result["_일반검색_섹션있음"] = general_exists
    # 2페이지로 실제로 이동했을 때만 통합검색 1페이지로 복귀
    if general_navigated:
        page.goto(integrated_url, wait_until="domcontentloaded", timeout=CURRENT_PAGE_LOAD_TIMEOUT_MS)
        page.wait_for_timeout(100)

    result["_t_일반 검색"] = time.perf_counter() - _t0
    result["_페이지이동됨"] = False
    return result


def _print_timing_summary(blog_id: str, keyword: str, t_goto: float, sections: dict, t_blogtab: float) -> None:
    label = f"{blog_id} / {keyword}" if blog_id else keyword
    line_w = max(len(label) + 6, 52)
    sep = "─" * line_w

    rows = [
        ("페이지 로드",     t_goto,                                None),
        ("검색결과",        sections.get("_t_검색결과", 0),        sections.get("검색결과")),
        ("반려동물 인기글", sections.get("_t_반려동물 인기글", 0), sections.get("반려동물 인기글")),
        ("일반 검색",       sections.get("_t_일반 검색", 0),       sections.get("일반 검색")),
        ("블로그(탭)",      t_blogtab,                             sections.get("블로그(탭)")),
    ]
    total = sum(t for _, t, _ in rows)

    print(f"\n━━ [{label}] " + "━" * max(0, line_w - len(label) - 5))
    for name, t, rank in rows:
        rank_str = f"→ {rank}위" if rank is not None else ("" if name == "페이지 로드" else "→ 없음")
        print(f"   {name:<16} {t:>5.1f}s  {rank_str}")
    print(f"   {sep}")
    print(f"   {'합계':<16} {total:>5.1f}s\n")


def check_naver_ranking(
    keyword: str,
    target_blog_id: str,
    headless: bool = True,
    use_debug_chrome: bool = False,
    debug_port: int = 9222,
) -> dict:
    """
    1) 통합검색(전체 탭) → 검색결과 / 반려동물 인기글 섹션에서 순위 수집
    2) 상단 lnb에서 '블로그' 탭 클릭 → 블로그 탭 페이지에서 순위 수집
    """
    target_id = normalize_blog_id(target_blog_id)
    encoded_kw = quote(keyword)
    integrated_url = f"https://search.naver.com/search.naver?query={encoded_kw}"
    out = {"keyword": keyword, "error": None, "sections": {}}

    with sync_playwright() as p:
        browser, context, cleanup = create_browser_session(
            p,
            headless=headless,
            use_debug_chrome=use_debug_chrome,
            debug_port=debug_port,
        )
        try:
            page = context.new_page()
            out = check_naver_ranking_on_page(page, keyword, target_blog_id)
        finally:
            cleanup()

    return out


def _scan_blog_tab_competitors(page, competitor_blog_ids: list[str]) -> dict[str, int | None]:
    """
    이미 로드된 블로그 탭 DOM에서 경쟁사 블로그ID별 순위를 재스캔한다.
    (추가 검색/네비게이션 없이 우리 블로그 스캔과 같은 컨테이너를 다시 훑는다.)
    """
    out: dict[str, int | None] = {}
    # 경쟁사는 우리보다 하위인 경우가 많아 상위 N개를 넉넉히 본다(블로그탭 로드분 한도 내).
    comp_max_cards = int(os.getenv("BLOG_COMPETITOR_MAX_CARDS", "30") or "30")
    for cid in competitor_blog_ids:
        nid = normalize_blog_id(cid)
        if not nid or nid in out:
            continue
        rank = None
        for container_sel in SELECTOR_BLOG_TAB_CONTAINER:
            r, _url = find_rank_in_cards(page, container_sel, CARD_BLOG_TAB, nid, max_cards=comp_max_cards)
            if r is not None:
                rank = r
                break
        out[nid] = rank
    return out


def check_naver_ranking_on_page(
    page, keyword: str, target_blog_id: str, competitor_blog_ids: list[str] | None = None
) -> dict:
    """
    주입된 Playwright page를 재사용하며 블로그 순위를 확인한다.
    (탭 누적 방지를 위해 page를 새로 만들지 않는다.)
    competitor_blog_ids 가 주어지면 같은 블로그 탭 로드 결과에서 경쟁사 순위도 함께 스캔한다.
    """
    target_id = normalize_blog_id(target_blog_id)
    encoded_kw = quote(keyword)
    integrated_url = f"https://search.naver.com/search.naver?query={encoded_kw}"
    out = {"keyword": keyword, "error": None, "sections": {}, "competitor_blog_tab": {}}

    for attempt in range(PAGE_LOAD_RETRY_COUNT + 1):
        timeout_ms = _set_page_load_timeout_for_attempt(attempt)
        phase = "통합검색 페이지 진입"
        attempt_label = f"{attempt + 1}/{PAGE_LOAD_RETRY_COUNT + 1}"
        print(f"🔎 [{target_id} / {keyword}] 시도 {attempt_label} 시작 (timeout={timeout_ms}ms)")
        try:
            print(f"   ↳ [{keyword}] {phase}")
            _t_goto = time.perf_counter()
            page.goto(integrated_url, wait_until="domcontentloaded", timeout=timeout_ms)
            page.wait_for_timeout(100)
            _t_goto = time.perf_counter() - _t_goto

            phase = "3개 섹션 순위 수집"
            print(f"   ↳ [{keyword}] {phase}")
            out["sections"] = get_three_section_ranks(page, target_id, integrated_url)

            phase = "블로그 탭 탐색"
            print(f"   ↳ [{keyword}] {phase}")
            _t_blogtab = time.perf_counter()
            blog_tab_loc = page.locator(SELECTOR_BLOG_TAB)
            if blog_tab_loc.count() > 0:
                phase = "블로그 탭 클릭/로드"
                print(f"   ↳ [{keyword}] {phase}")
                blog_tab_loc.first.click()
                page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
                page.wait_for_timeout(100)
                rank, blog_url = None, None
                for container_sel in SELECTOR_BLOG_TAB_CONTAINER:
                    rank, blog_url = find_rank_in_cards(
                        page, container_sel, CARD_BLOG_TAB, target_id, max_cards=20
                    )
                    if rank is not None:
                        break
                    more_button_sel = f'{container_sel} a[data-heatmap-target=".more"], {container_sel} a.more, {container_sel} button.more'
                    rank, blog_url = try_click_more_and_find_rank_blog_tab(
                        page, container_sel, CARD_BLOG_TAB, more_button_sel, target_id
                    )
                    if rank is not None:
                        break
                if rank is None:
                    rank, blog_url = _blog_tab_rank_fallback(page, target_id)
                out["sections"]["블로그(탭)"] = rank
                out["sections"]["블로그(탭)_URL"] = blog_url
                # 같은 블로그 탭 로드 결과에서 경쟁사 순위도 재스캔(추가 검색 없음)
                if competitor_blog_ids:
                    out["competitor_blog_tab"] = _scan_blog_tab_competitors(page, competitor_blog_ids)
            else:
                out["sections"]["블로그(탭)"] = None
                out["sections"]["블로그(탭)_URL"] = None
            _t_blogtab = time.perf_counter() - _t_blogtab

            out["error"] = None
            _print_timing_summary(target_id, keyword, _t_goto, out["sections"], _t_blogtab)
            print(f"✅ [{target_id} / {keyword}] 시도 {attempt_label} 완료")
            break
        except PlaywrightTimeout:
            out["error"] = "페이지 로드 시간 초과"
            print(f"⏱️ [{target_id} / {keyword}] 타임아웃 (단계: {phase}, 시도 {attempt_label})")
            if attempt >= PAGE_LOAD_RETRY_COUNT:
                break
        except Exception as e:
            out["error"] = str(e)
            print(f"❌ [{target_id} / {keyword}] 실패 (단계: {phase}, 시도 {attempt_label}): {e}")
            break

    return out


def read_input_excel(path: str) -> list[tuple[str, str]]:
    """엑셀에서 (블로그 ID, 키워드) 쌍 목록을 읽는다. 반드시 첫 번째 시트에서 읽음. A열=블로그 ID, B열=키워드. 빈 블로그 ID는 위 행 값 유지."""
    wb = openpyxl.load_workbook(path, read_only=False)
    ws = wb.worksheets[0]  # 항상 첫 번째 시트 (active가 아니라서 '플레이스' 시트가 선택돼 있어도 블로그는 1번 시트만 사용)
    pairs: list[tuple[str, str]] = []
    last_blog_id = ""
    for row in range(2, (ws.max_row or 0) + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        blog_id = (a and str(a).strip()) or last_blog_id
        keyword = (b and str(b).strip()) or ""
        if keyword and keyword.lower() != "키워드":
            if blog_id:
                last_blog_id = blog_id
                pairs.append((blog_id, keyword))
            # 블로그 ID가 비어 있고 last_blog_id도 없으면 첫 행이 헤더만 있는 경우라 스킵
    return pairs


def read_place_input_excel(path: str) -> list[tuple[str, str]]:
    """
    엑셀에서 플레이스 순위 확인용 (키워드, 상호명) 쌍 목록을 읽는다.
    시트 이름이 '플레이스'인 시트에서 A열=키워드, B열=상호명. 시트가 없으면 빈 목록 반환.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=False)
        if "플레이스" not in wb.sheetnames:
            return []
        ws = wb["플레이스"]
        pairs: list[tuple[str, str]] = []
        for row in range(2, (ws.max_row or 0) + 1):
            a = ws.cell(row, 1).value
            b = ws.cell(row, 2).value
            keyword = (a and str(a).strip()) or ""
            store_name = (b and str(b).strip()) or ""
            if keyword and keyword.lower() != "키워드" and store_name:
                pairs.append((keyword, store_name))
        return pairs
    except Exception:
        return []


def read_blog_input_from_supabase() -> list[tuple[str, str]]:
    """
    Supabase의 analytics.analytics_blog_keyword_targets 에서
    활성화된 (blog_id(account_id), keyword) 쌍을 읽는다.
    """
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("DB 입력을 사용하려면 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY가 필요합니다.")

    params = {
        "select": "account_id,keyword",
        "is_active": "eq.true",
        "order": "account_id.asc,keyword.asc",
    }
    target_hospital_id = os.getenv("COLLECT_HOSPITAL_ID", "").strip()
    if target_hospital_id:
        params["hospital_id"] = f"eq.{target_hospital_id}"
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_blog_keyword_targets?{urlencode(params)}"
    req = Request(url, headers=_supabase_headers(service_key, profile="analytics"), method="GET")
    with urlopen(req, timeout=20) as res:
        data = json.loads(res.read().decode("utf-8"))

    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in data or []:
        blog_id = str(row.get("account_id") or "").strip()
        keyword = str(row.get("keyword") or "").strip()
        if not blog_id or not keyword:
            continue
        key = (blog_id, keyword)
        if key in seen:
            continue
        seen.add(key)
        pairs.append(key)
    return pairs


def read_place_input_from_supabase() -> list[tuple[str, str, str | None]]:
    """
    Supabase의 analytics.analytics_place_keyword_targets + core.hospitals.name 에서
    활성화된 (keyword, store_name, hospital_id) 쌍을 읽는다.
    """
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        raise RuntimeError("DB 입력을 사용하려면 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY가 필요합니다.")

    params = {
        "select": "hospital_id,keyword",
        "is_active": "eq.true",
        "order": "hospital_id.asc,keyword.asc",
    }
    target_hospital_id = os.getenv("COLLECT_HOSPITAL_ID", "").strip()
    if target_hospital_id:
        params["hospital_id"] = f"eq.{target_hospital_id}"
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_place_keyword_targets?{urlencode(params)}"
    req = Request(url, headers=_supabase_headers(service_key, profile="analytics"), method="GET")
    with urlopen(req, timeout=20) as res:
        data = json.loads(res.read().decode("utf-8"))

    hospital_ids = sorted({str(row.get("hospital_id") or "").strip() for row in (data or []) if str(row.get("hospital_id") or "").strip()})
    hospital_name_map: dict[str, str] = {}
    if hospital_ids:
        in_values = ",".join(hospital_ids)
        hparams = {
            "select": "id,name",
            "id": f"in.({in_values})",
        }
        hurl = f"{supabase_url.rstrip('/')}/rest/v1/hospitals?{urlencode(hparams)}"
        hreq = Request(hurl, headers=_supabase_headers(service_key, profile="core"), method="GET")
        with urlopen(hreq, timeout=20) as res:
            hospitals = json.loads(res.read().decode("utf-8"))
        hospital_name_map = {
            str(row.get("id") or "").strip(): str(row.get("name") or "").strip()
            for row in (hospitals or [])
            if str(row.get("id") or "").strip()
        }

    pairs: list[tuple[str, str, str | None]] = []
    seen: set[tuple[str, str]] = set()
    for row in data or []:
        hospital_id = str(row.get("hospital_id") or "").strip()
        keyword = str(row.get("keyword") or "").strip()
        if not hospital_id or not keyword:
            continue
        store_name = hospital_name_map.get(hospital_id, "").strip()
        if not store_name:
            print(f"⚠️ 플레이스 입력 스킵: hospital_id={hospital_id}의 core.hospitals.name 이 비어 있습니다.")
            continue
        dedupe_key = (hospital_id, keyword)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        pairs.append((keyword, store_name, hospital_id))
    return pairs


# 병원별 경쟁사: {hospital_id: [(slot, name, blog_id), ...]} — main()에서 1회 로드, 워커가 참조.
_COMPETITORS_BY_HOSPITAL: dict[str, list[tuple[int, str, str | None]]] = {}


def read_competitors_from_supabase() -> dict[str, list[tuple[int, str, str | None]]]:
    """analytics_hospital_competitors 에서 활성 경쟁사 목록을 병원별로 읽는다.
    COLLECT_HOSPITAL_ID 가 있으면 그 병원만."""
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        return {}
    params = {
        "select": "hospital_id,slot,name,naver_blog_id",
        "is_active": "eq.true",
        "order": "hospital_id.asc,slot.asc",
    }
    target = os.getenv("COLLECT_HOSPITAL_ID", "").strip()
    if target:
        params["hospital_id"] = f"eq.{target}"
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_hospital_competitors?{urlencode(params)}"
    req = Request(url, headers=_supabase_headers(service_key, profile="analytics"), method="GET")
    out: dict[str, list[tuple[int, str, str | None]]] = {}
    try:
        with urlopen(req, timeout=20) as res:
            rows = json.loads(res.read().decode("utf-8"))
        for r in rows or []:
            hid = str(r.get("hospital_id") or "").strip()
            name = str(r.get("name") or "").strip()
            if not hid or not name:
                continue
            slot = int(r.get("slot") or 0)
            blog_id = str(r.get("naver_blog_id") or "").strip() or None
            out.setdefault(hid, []).append((slot, name, blog_id))
    except Exception as e:
        print(f"ℹ️ 경쟁사 조회 실패(무시): {e}")
    return out


# 블로그ID(account_id) → hospital_id 맵 — 블로그 워커가 경쟁사 조회용으로 참조.
_HOSPITAL_BY_BLOG_ID: dict[str, str] = {}


def read_blog_hospital_map_from_supabase() -> dict[str, str]:
    """analytics_blog_keyword_targets 에서 account_id → hospital_id 맵을 읽는다.
    COLLECT_HOSPITAL_ID 가 있으면 그 병원만. 블로그 워커가 경쟁사 매칭에 사용."""
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        return {}
    params = {
        "select": "account_id,hospital_id",
        "is_active": "eq.true",
    }
    target = os.getenv("COLLECT_HOSPITAL_ID", "").strip()
    if target:
        params["hospital_id"] = f"eq.{target}"
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_blog_keyword_targets?{urlencode(params)}"
    req = Request(url, headers=_supabase_headers(service_key, profile="analytics"), method="GET")
    out: dict[str, str] = {}
    try:
        with urlopen(req, timeout=20) as res:
            rows = json.loads(res.read().decode("utf-8"))
        for r in rows or []:
            bid = normalize_blog_id(str(r.get("account_id") or "").strip())
            hid = str(r.get("hospital_id") or "").strip()
            if bid and hid:
                out[bid] = hid
    except Exception as e:
        print(f"ℹ️ 블로그 hospital 맵 조회 실패(무시): {e}")
    return out


def write_output_excel(
    path: str, results: list[dict], place_results: list[dict] | None = None
) -> None:
    """결과 목록을 엑셀 파일로 저장한다. 순위에 들어온 경우 해당 섹션의 노출 URL 포함. place_results가 있으면 '플레이스 순위' 시트 추가."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "순위결과"
    ws.append([
        "블로그 ID", "키워드",
        "검색결과", "검색결과 노출URL",
        "반려동물 인기글", "반려동물 인기글 노출URL",
        "일반 검색", "일반 검색 노출URL",
        "블로그(탭)", "블로그(탭) 노출URL",
        "비고",
    ])

    def cell_val(sec: dict, key: str) -> str:
        rank = sec.get(key)
        if key == "검색결과" and rank is None:
            return "섹션 없음" if sec.get("_검색결과_섹션있음") is False else "—"
        if key == "반려동물 인기글" and rank is None:
            return "섹션 없음" if sec.get("_반려동물_섹션있음") is False else "—"
        if key == "일반 검색" and rank is None:
            return "섹션 없음" if sec.get("_일반검색_섹션있음") is False else "—"
        if rank is not None:
            return f"{rank}위"
        return "—"

    for r in results:
        blog_id = r.get("blog_id", "")
        kw = r.get("keyword", "")
        err = r.get("error") or ""
        sec = r.get("sections") or {}
        ws.append([
            blog_id,
            kw,
            cell_val(sec, "검색결과"),
            sec.get("검색결과_URL") or "",
            cell_val(sec, "반려동물 인기글"),
            sec.get("반려동물 인기글_URL") or "",
            cell_val(sec, "일반 검색"),
            sec.get("일반 검색_URL") or "",
            cell_val(sec, "블로그(탭)"),
            sec.get("블로그(탭)_URL") or "",
            err,
        ])

    if place_results:
        ws_place = wb.create_sheet("플레이스 순위")
        ws_place.append(["키워드", "상호명", "순위", "비고"])
        for r in place_results:
            rank = r.get("rank")
            ws_place.append([
                r.get("keyword", ""),
                r.get("store_name", ""),
                f"{rank}위" if rank is not None else "—",
                r.get("error") or "",
            ])
    wb.save(path)


def print_result(data: dict) -> None:
    """검색결과, 반려동물 인기글, 일반 검색, 블로그(탭) 네 곳 출력."""
    blog_id = data.get("blog_id", "")
    kw = data["keyword"]
    label = f"{blog_id} / {kw}" if blog_id else kw
    if data["error"]:
        print(f"⚠️ [{label}] 에러: {data['error']}\n")
        return

    sections = data["sections"]
    lines = [f"📌 [{label}]"]

    for key in ["검색결과", "반려동물 인기글", "일반 검색", "블로그(탭)"]:
        rank = sections.get(key)
        # 검색결과 섹션 없음 처리
        if key == "검색결과" and rank is None:
            if sections.get("_검색결과_섹션있음") == False:
                lines.append(f"   • {key}: 섹션 없음")
            else:
                lines.append(f"   • {key}: —")
        # 반려동물 인기글 섹션 없음 처리
        elif key == "반려동물 인기글" and rank is None:
            if sections.get("_반려동물_섹션있음") == False:
                lines.append(f"   • {key}: 섹션 없음")
            else:
                lines.append(f"   • {key}: —")
        # 일반 검색 섹션 없음 처리
        elif key == "일반 검색" and rank is None:
            if sections.get("_일반검색_섹션있음") == False:
                lines.append(f"   • {key}: 섹션 없음")
            else:
                lines.append(f"   • {key}: —")
        # 순위가 있으면 순위 + 노출 URL 표시
        elif rank is not None:
            lines.append(f"   • {key}: {rank}위")
            url = sections.get(
                "검색결과_URL" if key == "검색결과" else "반려동물 인기글_URL" if key == "반려동물 인기글" else "일반 검색_URL" if key == "일반 검색" else "블로그(탭)_URL"
            )
            if url:
                lines.append(f"      → {url}")
        else:
            lines.append(f"   • {key}: —")

    print("\n".join(lines) + "\n")


def _to_kst_date_str() -> str:
    now_utc = datetime.utcnow()
    kst = now_utc + timedelta(hours=9)
    return kst.strftime("%Y-%m-%d")


def _to_rank_value(rank) -> int | None:
    if rank is None:
        return None
    try:
        return int(rank)
    except Exception:
        return None


def _supabase_headers(service_key: str, profile: str | None = None) -> dict:
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
    }
    if profile:
        headers["Accept-Profile"] = profile
        headers["Content-Profile"] = profile
    return headers


def _check_rank_already_collected(supabase_url: str, service_key: str, hospital_id: str, metric_date: str) -> bool:
    """당일 해당 병원의 순위 데이터가 이미 존재하면 True."""
    if not hospital_id or not supabase_url or not service_key:
        return False
    params = {
        "select": "metric_date",
        "hospital_id": f"eq.{hospital_id}",
        "metric_date": f"eq.{metric_date}",
        "limit": "1",
    }
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_blog_keyword_ranks?{urlencode(params)}"
    try:
        req = Request(url, headers=_supabase_headers(service_key, profile="analytics"), method="GET")
        with urlopen(req, timeout=10) as res:
            data = json.loads(res.read().decode("utf-8"))
            return len(data) > 0
    except Exception:
        return False


def _supabase_get_hospital_map(supabase_url: str, service_key: str, blog_ids: list[str]) -> dict:
    cleaned = sorted({str(v).strip() for v in blog_ids if str(v).strip()})
    if not cleaned:
        return {}

    in_values = ",".join(cleaned)
    params = {
        "select": "id,name,naver_blog_id",
        "naver_blog_id": f"in.({in_values})",
    }
    url = f"{supabase_url.rstrip('/')}/rest/v1/hospitals?{urlencode(params)}"
    req = Request(url, headers=_supabase_headers(service_key, profile="core"), method="GET")
    with urlopen(req, timeout=20) as res:
        data = json.loads(res.read().decode("utf-8"))
    out = {}
    for row in data or []:
        blog_id = (row.get("naver_blog_id") or "").strip()
        if not blog_id:
            continue
        out[blog_id] = {
            "hospital_id": str(row.get("id")) if row.get("id") is not None else None,
            "hospital_name": row.get("name"),
        }
    return out


def _build_rank_upsert_payload(results: list[dict], hospital_map: dict, metric_date: str) -> list[dict]:
    payload = []
    collected_at = datetime.utcnow().isoformat() + "Z"
    for item in results:
        blog_id = str(item.get("blog_id") or "").strip()
        keyword = str(item.get("keyword") or "").strip()
        sections = item.get("sections") or {}
        error_msg = item.get("error")
        if not blog_id or not keyword:
            continue

        mapped = hospital_map.get(blog_id) or {}
        hospital_id = mapped.get("hospital_id")
        hospital_name = mapped.get("hospital_name")
        section_ranks: list[int] = []

        for label, url_key, section_key, metric_key in SECTION_SPECS:
            rank_num = _to_rank_value(sections.get(label))
            if rank_num is not None:
                section_ranks.append(rank_num)
                status = "found"
            else:
                if label == "검색결과" and sections.get("_검색결과_섹션있음") is False:
                    status = "section_missing"
                elif label == "반려동물 인기글" and sections.get("_반려동물_섹션있음") is False:
                    status = "section_missing"
                elif label == "일반 검색" and sections.get("_일반검색_섹션있음") is False:
                    status = "section_missing"
                else:
                    status = "not_found"

            payload.append({
                "account_id": blog_id,
                "hospital_id": hospital_id,
                "hospital_name": hospital_name,
                "source": "blog",
                "metric_date": metric_date,
                "metric_key": metric_key,
                "keyword": keyword,
                "section": section_key,
                "rank_value": rank_num,
                "exposed_url": sections.get(url_key),
                "metadata": {
                    "blog_id": blog_id,
                    "keyword": keyword,
                    "section_label": label,
                    "status": status,
                    "error": error_msg,
                },
                "collected_at": collected_at,
            })

        payload.append({
            "account_id": blog_id,
            "hospital_id": hospital_id,
            "hospital_name": hospital_name,
            "source": "blog",
            "metric_date": metric_date,
            "metric_key": "blog_rank_best",
            "keyword": keyword,
            "section": "최고 순위",
            "rank_value": min(section_ranks) if section_ranks else None,
            "exposed_url": None,
            "metadata": {
                "blog_id": blog_id,
                "keyword": keyword,
                "status": "found" if section_ranks else "not_found",
                "error": error_msg,
            },
            "collected_at": collected_at,
        })
    return payload


def upload_blog_ranks_to_supabase(results: list[dict], metric_date: str | None = None) -> int:
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        print("ℹ️ SUPABASE 환경변수가 없어 DB 업로드를 건너뜁니다.")
        return 0
    if not results:
        return 0

    resolved_metric_date = metric_date or os.getenv("RANK_METRIC_DATE") or _to_kst_date_str()
    hospital_map = _supabase_get_hospital_map(
        supabase_url,
        service_key,
        [str(r.get("blog_id") or "") for r in results],
    )
    payload = _build_rank_upsert_payload(results, hospital_map, resolved_metric_date)
    if not payload:
        return 0

    params = {
        "on_conflict": "account_id,metric_date,keyword,section,metric_key",
    }
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_blog_keyword_ranks?{urlencode(params)}"
    headers = _supabase_headers(service_key, profile="analytics")
    headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    req = Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    with urlopen(req, timeout=40):
        pass
    print(f"✅ Supabase 업서트 완료: {len(payload)}건 (metric_date={resolved_metric_date})")
    return len(payload)


def upload_place_ranks_to_supabase(place_results: list[dict], metric_date: str | None = None) -> int:
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        print("ℹ️ SUPABASE 환경변수가 없어 플레이스 DB 업로드를 건너뜁니다.")
        return 0
    if not place_results:
        return 0

    resolved_metric_date = metric_date or os.getenv("RANK_METRIC_DATE") or _to_kst_date_str()
    payload = []
    collected_at = datetime.utcnow().isoformat() + "Z"
    for item in place_results:
        keyword = str(item.get("keyword") or "").strip()
        store_name = str(item.get("store_name") or "").strip()
        if not keyword or not store_name:
            continue
        rank_num = _to_rank_value(item.get("rank"))
        error_msg = item.get("error")
        payload.append({
            "metric_date": resolved_metric_date,
            "hospital_id": str(item.get("hospital_id") or "").strip() or None,
            "keyword": keyword,
            "store_name": store_name,
            "section": "플레이스",
            "metric_key": "place_rank_integrated",
            "rank_value": rank_num,
            "metadata": {
                "status": "found" if rank_num is not None else "not_found",
                "error": error_msg,
            },
            "collected_at": collected_at,
        })

    if not payload:
        return 0

    params = {
        "on_conflict": "metric_date,keyword,store_name,section,metric_key",
    }
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_place_keyword_ranks?{urlencode(params)}"
    headers = _supabase_headers(service_key, profile="analytics")
    headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    req = Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    with urlopen(req, timeout=40):
        pass
    print(f"✅ Supabase 플레이스 업서트 완료: {len(payload)}건 (metric_date={resolved_metric_date})")
    return len(payload)


def upload_competitor_ranks_to_supabase(results: list[dict], channel: str, metric_date: str | None = None) -> int:
    """결과(item['competitors']=[{slot,name,rank}])에서 경쟁사 순위를 analytics_competitor_ranks 에 업서트."""
    supabase_url = os.getenv("SUPABASE_URL")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key or not results:
        return 0
    resolved_metric_date = metric_date or os.getenv("RANK_METRIC_DATE") or _to_kst_date_str()
    collected_at = datetime.utcnow().isoformat() + "Z"
    payload = []
    seen: set[tuple] = set()
    for item in results:
        hid = str(item.get("hospital_id") or "").strip()
        keyword = str(item.get("keyword") or "").strip()
        if not hid or not keyword:
            continue
        for c in item.get("competitors") or []:
            slot = int(c.get("slot") or 0)
            if slot < 1 or slot > 3:
                continue
            key = (hid, slot, channel, resolved_metric_date, keyword)
            if key in seen:
                continue
            seen.add(key)
            payload.append({
                "hospital_id": hid,
                "slot": slot,
                "channel": channel,
                "metric_date": resolved_metric_date,
                "keyword": keyword,
                "rank_value": _to_rank_value(c.get("rank")),
                "name": c.get("name"),
                "metadata": {},
                "collected_at": collected_at,
            })
    if not payload:
        return 0
    params = {"on_conflict": "hospital_id,slot,channel,metric_date,keyword"}
    url = f"{supabase_url.rstrip('/')}/rest/v1/analytics_competitor_ranks?{urlencode(params)}"
    headers = _supabase_headers(service_key, profile="analytics")
    headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    req = Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers, method="POST")
    with urlopen(req, timeout=40):
        pass
    print(f"✅ 경쟁사({channel}) 순위 업서트 완료: {len(payload)}건 (metric_date={resolved_metric_date})")
    return len(payload)


def _split_roundrobin(items: list, n: int) -> list[list]:
    """items를 n개 버킷에 라운드로빈으로 분배. 빈 버킷은 제거."""
    buckets: list[list] = [[] for _ in range(n)]
    for i, item in enumerate(items):
        buckets[i % n].append(item)
    return [b for b in buckets if b]


def _human_delay() -> None:
    """키워드 사이 사람처럼 보이는 랜덤 지연(고정 간격은 봇 신호). RANK_DELAY_MIN/MAX_SEC 로 조정."""
    lo = float(os.getenv("RANK_DELAY_MIN_SEC", "1.5"))
    hi = float(os.getenv("RANK_DELAY_MAX_SEC", "4.0"))
    if hi < lo:
        hi = lo
    time.sleep(random.uniform(lo, hi))


# 차단/캡차 감지 시 전 워커가 공유하는 정지 신호 (감지되면 그날 수집 중단 → IP 평판 보호)
_blocked_event = threading.Event()

_BLOCK_URL_MARKERS = ("captcha", "nid.naver.com/nidlogin", "/n/captcha", "blockmiddle")
_BLOCK_TEXT_MARKERS = (
    "자동 등록 방지", "비정상적", "보안문자", "캡차", "captcha",
    "일시적으로 제한", "접근이 제한",
)


def _detect_block(page) -> bool:
    """현재 페이지가 네이버 차단/캡차 페이지인지 판별.
    오탐 방지: URL 마커(captcha 리다이렉트 등)를 우선 신뢰하고, 본문 텍스트 마커는
    '정상 검색결과(#main_pack)가 없을 때'만 본다 — 검색결과 스니펫에 '비정상적'·'접근이 제한'
    같은 단어가 섞여 정상 페이지를 차단으로 오인하던 문제를 막는다."""
    try:
        url = (page.url or "").lower()
        for m in _BLOCK_URL_MARKERS:
            if m in url:
                print(f"🛑 차단 감지(URL 마커 '{m}'): {url[:120]}")
                return True
    except Exception:
        pass
    try:
        # 정상 검색결과/블로그탭 컨테이너(#main_pack)가 있으면 차단 페이지가 아니다.
        if page.query_selector("#main_pack"):
            return False
        low = (page.inner_text("body", timeout=1500) or "").lower()
        for m in _BLOCK_TEXT_MARKERS:
            if m.lower() in low:
                print(f"🛑 차단 감지(본문 마커 '{m}') — #main_pack 없음")
                return True
        return False
    except Exception:
        return False


def _probe_cdp(port: int) -> bool:
    """디버그 Chrome(CDP)이 해당 포트에 떠 있는지 확인."""
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=1.5) as resp:
            return getattr(resp, "status", 200) == 200
    except Exception:
        return False


_print_lock = threading.Lock()

# 키워드 순위 진행률(블로그+플레이스 합산). 워커 스레드들이 공유.
_progress_lock = threading.Lock()
_rank_progress = {"done": 0, "total": 0}

# 부분 결과 보존(증분 적재): 워커가 키워드 배치마다 즉시 DB 업서트할지 + 그날 metric_date + 배치 크기.
# 워커 타임아웃/네이버 차단/프로세스 강제종료가 나도 그때까지 수집한 순위는 DB에 남는다.
_PERSIST_INCREMENTAL = False
_PERSIST_METRIC_DATE: "str | None" = None
_PERSIST_BATCH = max(1, int(os.getenv("RANK_PERSIST_BATCH", "5")))

# 워커가 수집 즉시 여기에 쌓는다. future.result(timeout=) 이 터져도 이 버퍼는 살아남는다.
#
# 왜 필요한가: ThreadPoolExecutor 는 스레드를 죽이지 못한다. 타임아웃이 나도 `with` 블록을
# 빠져나갈 때 shutdown(wait=True) 로 워커가 끝날 때까지 어차피 기다린다. 즉 타임아웃은 시간을
# 아껴주지 못하면서 future 의 반환값만 버렸다. 그 결과 "블로그 0건 → 잡 실패" 오판이 났다.
# (실제로는 부분 적재로 DB 에 다 들어가 있었다.)
_shared_results_lock = threading.Lock()
_shared_blog_results: list = []
_shared_place_results: list = []


def _bump_rank_progress(label: str | None = None) -> None:
    with _progress_lock:
        _rank_progress["done"] += 1
        done = _rank_progress["done"]
        total = _rank_progress["total"]
    print(
        "__PROGRESS__ "
        + json.dumps({"step": "keyword_rank", "done": done, "total": total, "label": label}, separators=(",", ":")),
        flush=True,
    )


def _persist_blog_batch(rows: list[dict]) -> bool:
    """블로그 결과 배치를 즉시 업서트(부분 결과 보존). 성공 시 True(버퍼 비움), 실패 시 False(버퍼 유지→재시도)."""
    if not rows:
        return True
    try:
        upload_blog_ranks_to_supabase(rows, metric_date=_PERSIST_METRIC_DATE)
    except Exception as e:
        with _print_lock:
            print(f"⚠️ 블로그 부분 적재 실패(다음 배치에서 재시도): {e}")
        return False
    try:
        upload_competitor_ranks_to_supabase(rows, "blog", metric_date=_PERSIST_METRIC_DATE)
    except Exception as e:
        with _print_lock:
            print(f"⚠️ 경쟁사(블로그) 부분 적재 실패(무시): {e}")
    return True


def _persist_place_batch(rows: list[dict]) -> bool:
    """플레이스 결과 배치를 즉시 업서트(부분 결과 보존)."""
    if not rows:
        return True
    try:
        upload_place_ranks_to_supabase(rows, metric_date=_PERSIST_METRIC_DATE)
    except Exception as e:
        with _print_lock:
            print(f"⚠️ 플레이스 부분 적재 실패(다음 배치에서 재시도): {e}")
        return False
    try:
        upload_competitor_ranks_to_supabase(rows, "place", metric_date=_PERSIST_METRIC_DATE)
    except Exception as e:
        with _print_lock:
            print(f"⚠️ 경쟁사(플레이스) 부분 적재 실패(무시): {e}")
    return True


def _worker_blog_chunk(worker_id: int, pairs_chunk: list, use_debug_chrome: bool, debug_port: int) -> list[dict]:
    """
    워커 스레드: 자신만의 playwright + 브라우저 세션으로 pairs_chunk를 순차 처리.
    CDP 모드에서는 같은 Chrome 포트에 별도 연결하여 페이지만 신규 생성.
    """
    # Windows에서 스레드 내 asyncio ProactorEventLoop 충돌 방지
    if sys.platform == "win32":
        asyncio.set_event_loop(asyncio.SelectorEventLoop())
    chunk_results: list[dict] = []
    with sync_playwright() as p:
        _, context, cleanup = create_browser_session(
            p,
            headless=True,
            use_debug_chrome=use_debug_chrome,
            debug_port=debug_port,
        )
        page = context.new_page()
        pending: list[dict] = []
        try:
            for blog_id, kw in pairs_chunk:
                if _blocked_event.is_set():
                    break
                hospital_id = _HOSPITAL_BY_BLOG_ID.get(normalize_blog_id(blog_id), "")
                comps = _COMPETITORS_BY_HOSPITAL.get(hospital_id, []) if hospital_id else []
                competitor_blog_ids = [b for (_s, _n, b) in comps if b]
                data = check_naver_ranking_on_page(page, kw, blog_id, competitor_blog_ids)
                data["blog_id"] = blog_id
                cbt = data.get("competitor_blog_tab") or {}
                row = {
                    "blog_id": blog_id,
                    "keyword": kw,
                    "hospital_id": hospital_id or None,
                    "sections": data.get("sections"),
                    "error": data.get("error"),
                    "competitors": [
                        {"slot": slot, "name": name, "rank": cbt.get(normalize_blog_id(b))}
                        for (slot, name, b) in comps if b
                    ],
                }
                with _print_lock:
                    print_result(data)
                    if competitor_blog_ids:
                        comp_txt = ", ".join(
                            f"{name}:{cbt.get(normalize_blog_id(b))}" for (_s, name, b) in comps if b
                        )
                        print(f"   🏁 경쟁사 블로그(탭): {comp_txt}")
                chunk_results.append(row)
                with _shared_results_lock:
                    _shared_blog_results.append(row)
                _bump_rank_progress(kw)
                if _PERSIST_INCREMENTAL:
                    pending.append(row)
                    if len(pending) >= _PERSIST_BATCH and _persist_blog_batch(pending):
                        pending = []
                if _detect_block(page):
                    _blocked_event.set()
                    with _print_lock:
                        print(f"🛑 워커 {worker_id}: 네이버 차단/캡차 감지 — 수집 중단")
                    break
                _human_delay()
        finally:
            if _PERSIST_INCREMENTAL and pending:
                _persist_blog_batch(pending)
            with suppress(Exception):
                page.close()
            if not use_debug_chrome:
                with suppress(Exception):
                    cleanup()
    return chunk_results


def _worker_place_chunk(worker_id: int, place_chunk: list, use_debug_chrome: bool, debug_port: int) -> list[dict]:
    """
    워커 스레드: 자신만의 playwright + 브라우저 세션으로 place_chunk((kw, store, hospital_id))를 순차 처리.
    _worker_blog_chunk 와 동일 패턴.
    """
    if sys.platform == "win32":
        asyncio.set_event_loop(asyncio.SelectorEventLoop())
    chunk_results: list[dict] = []
    with sync_playwright() as p:
        _, context, cleanup = create_browser_session(
            p,
            headless=True,
            use_debug_chrome=use_debug_chrome,
            debug_port=debug_port,
        )
        page = context.new_page()
        pending: list[dict] = []
        try:
            for kw, store, hospital_id in place_chunk:
                if _blocked_event.is_set():
                    break
                comps = _COMPETITORS_BY_HOSPITAL.get(hospital_id or "", [])
                competitor_stores = [name for (_slot, name, _blog) in comps]
                data = check_place_ranking_on_page(page, kw, store, competitor_stores)
                data["hospital_id"] = hospital_id
                cr = data.get("competitor_ranks") or {}
                data["competitors"] = [
                    {"slot": slot, "name": name, "rank": cr.get(name)} for (slot, name, _blog) in comps
                ]
                with _print_lock:
                    err = data.get("error")
                    if err:
                        print(f"⚠️ [{kw} / {store}] 에러: {err}")
                    else:
                        comp_txt = ", ".join(f"{name}:{cr.get(name)}" for (_s, name, _b) in comps) if comps else ""
                        print(f"📌 [{kw} / {store}] 플레이스: {data.get('rank')}위" + (f" | 경쟁사 {comp_txt}" if comp_txt else ""))
                chunk_results.append(data)
                with _shared_results_lock:
                    _shared_place_results.append(data)
                _bump_rank_progress(f"{kw} / {store}")
                if _PERSIST_INCREMENTAL:
                    pending.append(data)
                    if len(pending) >= _PERSIST_BATCH and _persist_place_batch(pending):
                        pending = []
                if _detect_block(page):
                    _blocked_event.set()
                    with _print_lock:
                        print(f"🛑 워커 {worker_id}: 네이버 차단/캡차 감지 — 수집 중단")
                    break
                _human_delay()
        finally:
            if _PERSIST_INCREMENTAL and pending:
                _persist_place_batch(pending)
            with suppress(Exception):
                page.close()
            if not use_debug_chrome:
                with suppress(Exception):
                    cleanup()
    return chunk_results


def main():
    print(f"▶ naver-rank-main.py 시작 (Python {sys.version.split()[0]}, platform={sys.platform})")
    input_path = "input.xlsx"
    output_path = "output.xlsx"
    export_excel = False
    upload_db = True
    metric_date = None
    # 순위 수집(블로그/플레이스)은 로그인이 필요 없다 → 병원별(로그인) Chrome 포트
    # (CHROME_DEBUGGING_PORT)는 절대 쓰지 않고, 항상 비로그인 전용 포트로 간다.
    # 기본 9223 = scripts/windows/chrome-debug-rank-port9223.cmd 가 띄우는 순위 전용 비로그인 크롬.
    # 다른 비로그인 포트를 쓰고 싶을 때만 RANK_CHROME_DEBUGGING_PORT 로 override.
    # (계정/봇 탐지 리스크를 없애기 위함 — 병원 포트로 폴백하던 동작을 의도적으로 제거.)
    rank_port_raw = os.getenv("RANK_CHROME_DEBUGGING_PORT")
    raw_debug_port = rank_port_raw or "9223"
    input_source = os.getenv("RANK_INPUT_SOURCE", "db").strip().lower()
    try:
        debug_port = int(raw_debug_port)
    except ValueError:
        print(f"❌ Chrome 디버깅 포트 값이 잘못되었습니다: {raw_debug_port}")
        return
    if rank_port_raw:
        print(f"ℹ️ 순위 전용 Chrome 포트 사용: RANK_CHROME_DEBUGGING_PORT={debug_port}")

    # CDP(실 Chrome) 가 가장 탐지에 안전 → 기본은 'CDP 전용'. 헤드리스 자동 폴백은 하지 않는다.
    # (조용히 더 잘 들키는 모드로 떨어지거나, 좋은 데이터를 not_found 로 덮어쓰는 사고를 막기 위함.)
    # 로컬 개발 등에서 독립 Chromium 으로 돌리려면 RANK_USE_DEBUG_CHROME=0 으로 명시적으로 끈다.
    _explicit_cdp = os.getenv("RANK_USE_DEBUG_CHROME")
    if _explicit_cdp is not None:
        use_debug_chrome = _is_truthy_env("RANK_USE_DEBUG_CHROME")
    else:
        use_debug_chrome = True  # 기본 CDP 전용(폴백 없음)
    if use_debug_chrome:
        # 디버그 Chrome 이 안 떠 있으면 headless 로 떨어지지 않고 '시끄럽게' 실패(잡 failed)한다.
        if not _probe_cdp(debug_port):
            print(f"❌ 디버그 Chrome(CDP)이 port={debug_port} 에 없습니다. "
                  f"Chrome 을 --remote-debugging-port={debug_port} 로 실행한 뒤 다시 시도하세요. "
                  f"(headless 폴백 비활성화 — 독립 Chromium 으로 강제하려면 RANK_USE_DEBUG_CHROME=0)")
            sys.exit(1)
        print(f"ℹ️ CDP 모드(실 Chrome, port={debug_port}) 사용 — 탐지 위험 최소")
    else:
        print("ℹ️ 독립 Chromium launch 모드 (RANK_USE_DEBUG_CHROME=0 명시).")

    positional = []
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        arg = args[i]
        if arg == "--db-only":
            # Backward compatibility: --db-only means do not write excel.
            export_excel = False
            i += 1
            continue
        if arg == "--no-db":
            upload_db = False
            i += 1
            continue
        if arg == "--export-excel":
            export_excel = True
            i += 1
            continue
        if arg == "--metric-date":
            if i + 1 >= len(args):
                print("❌ --metric-date 다음에 YYYY-MM-DD를 입력하세요.")
                return
            metric_date = args[i + 1]
            i += 2
            continue
        if arg == "--use-debug-chrome":
            use_debug_chrome = True
            i += 1
            continue
        if arg == "--debug-port":
            if i + 1 >= len(args):
                print("❌ --debug-port 다음에 포트 번호를 입력하세요.")
                return
            try:
                debug_port = int(args[i + 1])
            except ValueError:
                print("❌ --debug-port는 숫자여야 합니다. 예: --debug-port 9222")
                return
            i += 2
            continue
        if arg == "--input-source":
            if i + 1 >= len(args):
                print("❌ --input-source 다음에 db 또는 excel 을 입력하세요.")
                return
            input_source = args[i + 1].strip().lower()
            if input_source not in {"db", "excel"}:
                print("❌ --input-source는 db 또는 excel 이어야 합니다.")
                return
            i += 2
            continue
        positional.append(arg)
        i += 1

    if len(positional) >= 1:
        input_path = positional[0]
    if len(positional) >= 2:
        output_path = positional[1]

    if input_source == "db":
        try:
            pairs = read_blog_input_from_supabase()
            place_pairs = read_place_input_from_supabase()
        except Exception as e:
            print(f"❌ DB 입력 조회 실패: {e}")
            return
        # 경쟁사 로드(1회) — 워커가 같은 검색에서 우리+경쟁사 순위를 함께 찾는다.
        global _COMPETITORS_BY_HOSPITAL, _HOSPITAL_BY_BLOG_ID
        _COMPETITORS_BY_HOSPITAL = read_competitors_from_supabase()
        _HOSPITAL_BY_BLOG_ID = read_blog_hospital_map_from_supabase()
        if _COMPETITORS_BY_HOSPITAL:
            _n = sum(len(v) for v in _COMPETITORS_BY_HOSPITAL.values())
            print(f"🏁 경쟁사 {_n}곳 로드({len(_COMPETITORS_BY_HOSPITAL)}개 병원) — 순위 비교 대상 포함")
    else:
        if Path(input_path).exists():
            pairs = read_input_excel(input_path)
        else:
            target = TARGET_BLOG_ID
            pairs = [(target, kw) for kw in KEYWORDS] if target and KEYWORDS else []
        place_pairs = [(kw, store, None) for kw, store in (read_place_input_excel(input_path) if Path(input_path).exists() else [])]

    # 플레이스 순위만 빠르게 보고 싶을 때(블로그가 먼저 다 돌아 오래 걸림) RANK_SKIP_BLOG=1.
    if _is_truthy_env("RANK_SKIP_BLOG") and pairs:
        print(f"⏭️ RANK_SKIP_BLOG=1 — 블로그 순위 {len(pairs)}개 건너뜀(플레이스만 수집)")
        pairs = []
    # 반대로 블로그만 보려면 RANK_SKIP_PLACE=1.
    if _is_truthy_env("RANK_SKIP_PLACE") and place_pairs:
        print(f"⏭️ RANK_SKIP_PLACE=1 — 플레이스 순위 {len(place_pairs)}개 건너뜀(블로그만 수집)")
        place_pairs = []

    if not pairs and not place_pairs:
        if input_source == "db":
            print("❌ 활성화된 키워드 타깃이 없습니다. analytics.analytics_blog_keyword_targets를 확인하세요.")
        else:
            print("❌ 확인할 (블로그 ID, 키워드) 또는 (키워드, 상호명)이 없습니다. input.xlsx를 확인하세요.")
        return
    print(f"ℹ️ 입력 소스: {input_source}")

    supabase_url = os.getenv("SUPABASE_URL", "")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
    target_hospital_id = os.getenv("COLLECT_HOSPITAL_ID", "").strip()
    check_date = metric_date or os.getenv("RANK_METRIC_DATE") or _to_kst_date_str()
    # 기본: 같은 날 재수집하면 최신 값으로 덮어쓴다(날짜 키 upsert merge-duplicates).
    # 예전처럼 "그날 이미 수집됐으면 스킵"하려면 RANK_SKIP_IF_COLLECTED=1 로 켠다.
    if _is_truthy_env("RANK_SKIP_IF_COLLECTED") and target_hospital_id and supabase_url and service_key:
        if _check_rank_already_collected(supabase_url, service_key, target_hospital_id, check_date):
            print(f"ℹ️ 키워드 순위 이미 수집됨 (hospital_id={target_hospital_id}, date={check_date}). 스킵합니다.")
            return

    num_workers = int(os.getenv("RANK_PARALLEL_WORKERS", "1"))
    print(f"ℹ️ 병렬 워커 수: {num_workers} (RANK_PARALLEL_WORKERS)")
    results: list[dict] = []
    place_results: list[dict] = []

    # 진행률 total = 블로그 유효 조합 + 플레이스 조합 (워커들이 done 증가)
    _blog_valid_count = len([(b, k) for b, k in pairs if b and b != "your_blog_id"]) if pairs else 0
    with _progress_lock:
        _rank_progress["done"] = 0
        _rank_progress["total"] = _blog_valid_count + (len(place_pairs) if place_pairs else 0)

    # 부분 결과 보존: 워커가 키워드 배치마다 즉시 DB 적재(타임아웃/차단/강제종료에도 수집분 유지).
    global _PERSIST_INCREMENTAL, _PERSIST_METRIC_DATE
    _PERSIST_INCREMENTAL = bool(upload_db)
    _PERSIST_METRIC_DATE = metric_date
    if _PERSIST_INCREMENTAL:
        print(f"ℹ️ 부분 결과 보존 ON — {_PERSIST_BATCH}개 단위로 즉시 적재")

    # 블로그 키워드: 병렬 처리 (워커별 독립 playwright 세션 + 페이지)
    if pairs:
        valid_pairs = [(b, k) for b, k in pairs if b and b != "your_blog_id"]
        skipped = len(pairs) - len(valid_pairs)
        if skipped:
            print(f"⚠️ 블로그 ID 없는 {skipped}개 키워드 스킵")
        chunks = _split_roundrobin(valid_pairs, num_workers)
        actual_workers = len(chunks)
        worker_timeout_sec = int(os.getenv("RANK_WORKER_TIMEOUT_SEC", "3000"))
        print(f"🚀 네이버 블로그 순위 확인 — {len(valid_pairs)}개 조합 / {actual_workers}개 병렬 워커 (워커 타임아웃: {worker_timeout_sec}s)\n")
        with ThreadPoolExecutor(max_workers=actual_workers) as executor:
            futures = [
                executor.submit(_worker_blog_chunk, i, chunk, use_debug_chrome, debug_port)
                for i, chunk in enumerate(chunks)
            ]
            for i, future in enumerate(futures):
                try:
                    future.result(timeout=worker_timeout_sec)
                except FutureTimeoutError:
                    # 스레드는 계속 돈다(파이썬은 스레드를 죽일 수 없다). 아래 shutdown(wait=True) 에서
                    # 어차피 끝날 때까지 기다리므로, 수집분은 공유 버퍼에서 그대로 살아남는다.
                    print(f"⚠️ 워커 {i} 타임아웃 ({worker_timeout_sec}s 초과) — 계속 진행, 수집분은 유지")
                except Exception as e:
                    print(f"❌ 워커 {i} 실패: {e}")
        # 워커가 실제로 모은 것을 쓴다. future 반환값이 아니라 공유 버퍼가 진실이다.
        with _shared_results_lock:
            results.extend(_shared_blog_results)

    # 플레이스 키워드: 병렬 처리 (블로그와 동일 패턴 — 워커별 독립 세션)
    if place_pairs:
        place_chunks = _split_roundrobin(place_pairs, num_workers)
        actual_place_workers = len(place_chunks)
        worker_timeout_sec = int(os.getenv("RANK_WORKER_TIMEOUT_SEC", "3000"))
        print(f"\n🏪 플레이스 순위 확인 — {len(place_pairs)}개 조합 / {actual_place_workers}개 병렬 워커 (광고 제외)\n")
        with ThreadPoolExecutor(max_workers=actual_place_workers) as executor:
            futures = [
                executor.submit(_worker_place_chunk, i, chunk, use_debug_chrome, debug_port)
                for i, chunk in enumerate(place_chunks)
            ]
            for i, future in enumerate(futures):
                try:
                    future.result(timeout=worker_timeout_sec)
                except FutureTimeoutError:
                    print(f"⚠️ 플레이스 워커 {i} 타임아웃 ({worker_timeout_sec}s 초과) — 계속 진행, 수집분은 유지")
                except Exception as e:
                    print(f"❌ 플레이스 워커 {i} 실패: {e}")
        with _shared_results_lock:
            place_results.extend(_shared_place_results)

    if upload_db and results:
        try:
            upload_blog_ranks_to_supabase(results, metric_date=metric_date)
        except Exception as e:
            print(f"❌ Supabase 업로드 실패: {e}")
        try:
            upload_competitor_ranks_to_supabase(results, "blog", metric_date=metric_date)
        except Exception as e:
            print(f"❌ 경쟁사(블로그) 순위 업로드 실패: {e}")
    if upload_db and place_results:
        try:
            upload_place_ranks_to_supabase(place_results, metric_date=metric_date)
        except Exception as e:
            print(f"❌ Supabase 플레이스 업로드 실패: {e}")
        try:
            upload_competitor_ranks_to_supabase(place_results, "place", metric_date=metric_date)
        except Exception as e:
            print(f"❌ 경쟁사(플레이스) 순위 업로드 실패: {e}")

    if export_excel:
        write_output_excel(output_path, results, place_results if place_results else None)
        print(f"\n✅ 결과 저장: {output_path}")
    else:
        print("\nℹ️ 기본 모드(DB 직적재)로 엑셀 저장을 건너뜁니다. 필요 시 --export-excel 사용")
    if _blocked_event.is_set():
        print("🛑 네이버 차단/캡차 감지로 수집을 중단했습니다 — 일부만 수집됨. "
              "잠시 후(가능하면 IP/시간대 변경) 재시도하세요.")

    # ★가짜 성공 방지: 수집할 타깃이 있었는데 한 건도 못 모았으면 성공(exit 0)으로 끝내지 않는다.
    #  판단 근거는 공유 버퍼(워커가 실제로 모은 것)다. 예전엔 future 반환값을 봤는데, 워커 타임아웃이
    #  그 값을 버려서 '멀쩡히 수집된 잡'이 실패로 찍혔다. 0건은 이제 진짜 0건이다.
    #  원인을 CDP 연결 실패로 단정하지 않는다 — 타임아웃·차단·크롬 부재 모두 여기로 온다.
    blog_expected = _blog_valid_count > 0
    place_expected = bool(place_pairs)
    if (blog_expected and not results) or (place_expected and not place_results):
        failed = []
        if blog_expected and not results:
            failed.append("블로그")
        if place_expected and not place_results:
            failed.append("플레이스")
        print(
            f"❌ {'/'.join(failed)} 순위를 한 건도 수집하지 못했습니다. "
            f"위 로그에서 워커 실패 사유(디버그 Chrome(CDP) 연결·타임아웃·네이버 차단)를 확인하세요. "
            f"잡을 실패 처리합니다.",
            file=sys.stderr,
        )
        sys.exit(1)

    print("✅ 끝.")


if __name__ == "__main__":
    main()
